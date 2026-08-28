'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃  🌟 Empresa: Ionclinics                                                                                                                                    ┃
┃  🚀 Proyecto: Auto Bach Record                                                                                                                             ┃
┃  🌈 Versión: v0.0                                                                                                                                          ┃
┃  👨‍💻 Desarrollador: Miguel Sánchez Faubel                                                                                                                   ┃        
┣━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┫
┃  📝 Descripción: Automatización del BachRecord                                                                                                             ┃      
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
'''
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ IMPORTACION Y CONFIGURACION DE LIBRERIAS ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

import os
from duckdb import df
import pandas as pd
import time
import re
from pathlib import Path
import unicodedata
import pythoncom
import win32com.client as win32
import win32print
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side
import pickle
from pprint import pprint
from pdf_gen import *

import winreg
import win32com.client
import win32print

from pypdf import PdfReader, PdfWriter, PageObject, Transformation
from reportlab.lib.pagesizes import A4

import os

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ "COMPILACION" ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛

MODO_RESET = True # Este modo recalcula todos los resumenes a partir de las ordenes de produccion desde 0
GENERAR_PDF_ORDENES = True # True genera las ordenes, false no las genera

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ CONSTANTES ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛

CONSOLAS = ['EPTEV01V01_CON','EPTEV01V02_CON','EPTEV02V01_CON','EPTEV02V02_CON'] # ARTICULOS QUE NO SE BUSCA EL PNT EN EL RESUMEN DEL ARTICULO
DISPOSITIVOS = ['EPTEV02DEV01', 'EPTEV02DEV02', 'EPTEV01DEV01', 'EPTEV01DEV02']
INDICE_ARTICULOS = ['MATERIA PRIMA RAW', 'MATERIA PRIMA N1', 'MATERIA PRIMA N2', 'DISPOSITIVOS'] # SIEMPRE ORDENADOS DE MAS PROCESADOS A MENOS PROCESADOS
DISPOSITIVO_CONSOLA = {'EPTEV01DEV01':'EPTEV01V01_CON','EPTEV01DEV02': 'EPTEV01V02_CON','EPTEV02DEV01': 'EPTEV02V01_CON','EPTEV02DEV02': 'EPTEV02V02_CON'}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ COMPROBAMOS LOS DIRECTORIOS ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛

RUTA_RAIZ = './data/'
RUTA_ERP =  RUTA_RAIZ + 'PNT MDR/ERP/'
RUTA_IMD = RUTA_RAIZ + 'PNT MDR/ERP/I+D/'
RUTA_PNT =  RUTA_RAIZ +'PNT MDR/PNT.7.5-02_CONTROL DE CALIDAD/Registros MDR'
RUTA_PNT_EXCEL = RUTA_RAIZ +'PNT MDR/PNT.7.5-02_CONTROL DE CALIDAD/Registros MDR/REG.7.5-02-02_CONTROL PNT.xlsx'
RUTA_CONFIG ='./config/config.xlsx'
RUTA_ETIQUETAS = './data/IMPRESION_ETIQUETAS/' # PARA LOS REQUISITOS ESPECIALES DEL BACH RECORD DISPOSITIVOS FUERA DE ESPAÑA
RUTA_PDF_BLANCO = './config/hoja_en_blanco.pdf'

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ GLOBALES ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

lista_errores = []

patron_lotes = re.compile(
    r"^(?P<articulo>.*?)\s*"
    r"\(\s*"
    r"(?P<serie_inicial>[A-Za-z0-9_]+)"
    r"\s*-\s*"
    r"(?P<serie_final>[A-Za-z0-9_]+)"
    r"\s*\)"
)

patron_orden_produccion = re.compile(
    r"^\d{3,4} \d{2}_\d{2}_\d{2}$"
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ FUNCIONES ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛

def convertir_excels_lotes_erp_a_pdf(carpeta_erp=RUTA_ERP):

    '''
    COMBIERTE A PDF LOS LOTES DE LOS ARTICULOS COMPRADOS
    '''
    carpeta_erp = Path(carpeta_erp)
    extensiones_excel = {".xlsx", ".xls", ".xlsm", ".xlsb"}

    pdf_generados = []
    errores = []

    if not carpeta_erp.exists():
        raise FileNotFoundError(f"No existe la carpeta ERP: {carpeta_erp}")

    for raiz, carpetas, archivos in os.walk(carpeta_erp):

        # No entrar en I+D
        carpetas[:] = [carpeta for carpeta in carpetas if carpeta.strip().upper() != "I+D"]

        for nombre_archivo in archivos:

            ruta_excel = Path(raiz) / nombre_archivo

            if ruta_excel.suffix.lower() not in extensiones_excel:
                continue

            if ruta_excel.name.startswith("~$"):
                continue

            if "LOTES" not in ruta_excel.stem.upper():
                continue
            if "_FORMATEADO" not in ruta_excel.stem.upper():
                continue

            try:
                print(f"Convirtiendo: {ruta_excel}")

                ruta_pdf = excel_resumen_a_pdf_una_hoja(ruta_excel)

                pdf_generados.append(str(ruta_pdf))

            except Exception as e:
                errores.append({"excel": str(ruta_excel), "error": str(e)})
                print(f"ERROR convirtiendo {ruta_excel}: {e}")

    print(f"\nPDF generados: {len(pdf_generados)}")
    print(f"Errores: {len(errores)}")

    return pdf_generados, errores

def guardar_diccionario(diccionario, ruta="./temp/estructura_documental.pkl"):
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    with open(ruta, "wb") as archivo:
        pickle.dump(diccionario, archivo)

    return str(ruta)

def cargar_diccionario(ruta="./temp/estructura_documental.pkl"):
    with open(ruta, "rb") as archivo:
        return pickle.load(archivo)

def eliminar_articulos_vacios_df(df):
    """
    Elimina bloques completos de artículos cuando TODAS sus columnas
    están completamente vacías.

    Estructura esperada por artículo:

        ARTICULO_LOTE
        ARTICULO_UNID
        ARTICULO_ALBARAN
        ARTICULO_PNT

    Ejemplo:

        PDB1V01_LOTE
        PDB1V01_UNID
        PDB1V01_ALBARAN
        PDB1V01_PNT

    Si las cuatro están vacías -> elimina las cuatro.
    Si cualquiera tiene información -> conserva las cuatro.
    """

    df = df.copy()

    # ========================================================
    # QUÉ CONSIDERAMOS VACÍO
    # ========================================================

    def es_vacio(valor):

        # NaN, None, pd.NA...
        if pd.isna(valor):
            return True

        # Cadenas vacías o espacios
        if isinstance(valor, str):
            return valor.strip() == ""

        return False


    # ========================================================
    # IDENTIFICAR COLUMNAS POR ARTÍCULO
    # ========================================================

    sufijos = [
        "_LOTE",
        "_UNID",
        "_ALBARAN",
        "_PNT"
    ]

    articulos = {}


    for columna in df.columns:

        nombre = str(columna).strip()

        nombre_upper = nombre.upper()

        for sufijo in sufijos:

            if nombre_upper.endswith(sufijo):

                # Ejemplo:
                #
                # PDB1V01_LOTE
                #      ↓
                # PDB1V01

                articulo = nombre[
                    :-len(sufijo)
                ]

                if articulo not in articulos:

                    articulos[articulo] = []

                articulos[articulo].append(
                    columna
                )

                break


    # ========================================================
    # COMPROBAR CADA BLOQUE
    # ========================================================

    columnas_eliminar = []


    for articulo, columnas_articulo in articulos.items():

        # ----------------------------------------------------
        # Comprobar si TODAS las celdas de TODAS las columnas
        # del artículo están vacías
        # ----------------------------------------------------

        bloque_vacio = True


        for columna in columnas_articulo:

            for valor in df[columna]:

                if not es_vacio(valor):

                    bloque_vacio = False
                    break

            if not bloque_vacio:
                break


        # ----------------------------------------------------
        # Si todo está vacío, eliminar el bloque completo
        # ----------------------------------------------------

        if bloque_vacio:

            columnas_eliminar.extend(columnas_articulo)
            # print(f"Artículo vacío eliminado: {articulo}")


    # ========================================================
    # ELIMINAR COLUMNAS
    # ========================================================

    df = df.drop(columns=columnas_eliminar)

    return df

def obtener_unico_excel(directorio):
    """
    Busca archivos Excel en un directorio y comprueba
    que exista exactamente uno.

    Devuelve:
        Path del Excel encontrado.

    Genera:
        FileNotFoundError si no hay ningún Excel.
        ValueError si hay más de un Excel.
    """

    directorio = Path(directorio)

    if not directorio.exists():
        raise FileNotFoundError(
            f"El directorio no existe: {directorio}"
        )

    if not directorio.is_dir():
        raise NotADirectoryError(
            f"La ruta no es un directorio: {directorio}"
        )

    extensiones_excel = {
        ".xlsx",
        ".xlsm",
        ".xls"
    }

    excels = [
        archivo
        for archivo in directorio.iterdir()
        if archivo.is_file()
        and archivo.suffix.lower() in extensiones_excel
        and not archivo.name.startswith("~$")
    ]

    # Ordenamos por nombre
    excels.sort()

    if len(excels) == 0:
        raise FileNotFoundError(
            f"No se ha encontrado ningún Excel en: {directorio}"
        )

    if len(excels) > 1:
        nombres = "\n".join(
            f" - {archivo.name}"
            for archivo in excels
        )

        raise ValueError(
            f"Se esperaba un único Excel, pero se han encontrado "
            f"{len(excels)}:\n{nombres}"
        )

    return excels[0]

def excel_resumen_a_pdf_una_hoja(ruta_excel, ruta_pdf=None, hoja=None, horizontal=True):
    ruta_excel = Path(ruta_excel).resolve()

    if not ruta_excel.exists():
        raise FileNotFoundError(f"No existe el Excel: {ruta_excel}")

    if ruta_pdf is None:
        ruta_pdf = ruta_excel.with_suffix(".pdf")
    else:
        ruta_pdf = Path(ruta_pdf).resolve()

    ruta_pdf.parent.mkdir(parents=True, exist_ok=True)

    excel = None
    libro = None
    impresora_original = None

    try:
        # ========================================================
        # GUARDAR IMPRESORA ACTUAL Y UTILIZAR MICROSOFT PRINT TO PDF
        # ========================================================

        impresora_original = win32print.GetDefaultPrinter()

        impresoras = [impresora[2] for impresora in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)]

        impresora_pdf = next((nombre for nombre in impresoras if "MICROSOFT PRINT TO PDF" in nombre.upper()), None)

        if impresora_pdf:
            win32print.SetDefaultPrinter(impresora_pdf)

        # ========================================================
        # ABRIR EXCEL
        # ========================================================

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # ========================================================
        # PRIMERO ABRIR EL LIBRO
        # ========================================================

        libro = excel.Workbooks.Open(str(ruta_excel))

        if hoja is None:
            ws = libro.Worksheets(1)
        else:
            ws = libro.Worksheets(hoja)

        ws.Activate()
        ruta_registro = r"Software\Microsoft\Windows NT\CurrentVersion\Devices"
        nombre_impresora_pdf = "Microsoft Print to PDF"

        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, ruta_registro) as clave:
            valor, _ = winreg.QueryValueEx(clave, nombre_impresora_pdf)

        puerto_excel = valor.split(",")[-1].strip()

        if not puerto_excel.endswith(":"):
            puerto_excel += ":"

        impresora_excel = f"{nombre_impresora_pdf} en {puerto_excel}"

        print(f"IMPRESORA ACTUAL EXCEL: {excel.ActivePrinter}")
        print(f"Intentando configurar Excel con: {impresora_excel}")

        excel.ActivePrinter = impresora_excel

        print(f"IMPRESORA ACTIVA EN EXCEL: {excel.ActivePrinter}")
        # ========================================================
        # ÚLTIMA FILA Y COLUMNA CON INFORMACIÓN
        # ========================================================

        # ultima_celda_fila = ws.Cells.Find(What="*", After=ws.Cells(1, 1), LookIn=-4123, SearchOrder=1, SearchDirection=2)
        # ultima_celda_columna = ws.Cells.Find(What="*", After=ws.Cells(1, 1), LookIn=-4123, SearchOrder=2, SearchDirection=2)

        # if ultima_celda_fila is None or ultima_celda_columna is None:
        #     raise ValueError(f"La hoja '{ws.Name}' no contiene información.")
        
        rango_usado = ws.UsedRange

        ultima_fila = rango_usado.Row + rango_usado.Rows.Count - 1
        ultima_columna = rango_usado.Column + rango_usado.Columns.Count - 1

        print(f"ULTIMA FILA: {ultima_fila}")
        print(f"ULTIMA COLUMNA: {ultima_columna}")
        print(f"RANGO: A1:{ws.Cells(ultima_fila, ultima_columna).Address.replace('$', '')}")

        rango = ws.Range(ws.Cells(1, 1), ws.Cells(ultima_fila, ultima_columna))

        # ========================================================
        # EXPORTAR
        # ========================================================
        
        ws.Activate()
        ws.ResetAllPageBreaks()

        excel.PrintCommunication = False

        ws.PageSetup.PrintArea = rango.Address
        ws.PageSetup.Orientation = 2 if horizontal else 1
        ws.PageSetup.PaperSize = 9

        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = 1

        ws.PageSetup.LeftMargin = excel.CentimetersToPoints(0.5)
        ws.PageSetup.RightMargin = excel.CentimetersToPoints(0.5)
        ws.PageSetup.TopMargin = excel.CentimetersToPoints(0.5)
        ws.PageSetup.BottomMargin = excel.CentimetersToPoints(0.5)
        ws.PageSetup.HeaderMargin = 0
        ws.PageSetup.FooterMargin = 0

        ws.PageSetup.CenterHorizontally = True
        ws.PageSetup.CenterVertically = True

        excel.PrintCommunication = True

        print(f"Zoom: {ws.PageSetup.Zoom}")
        print(f"Fit ancho: {ws.PageSetup.FitToPagesWide}")
        print(f"Fit alto: {ws.PageSetup.FitToPagesTall}")
        print(f"Área impresión: {ws.PageSetup.PrintArea}")

        ws.ExportAsFixedFormat(0, str(ruta_pdf), Quality=0, IncludeDocProperties=True, IgnorePrintAreas=False, OpenAfterPublish=False)

        return str(ruta_pdf)

    finally:
        if libro is not None:
            libro.Close(SaveChanges=False)

        if excel is not None:
            excel.Quit()

        # ========================================================
        # RESTAURAR IMPRESORA ORIGINAL
        # ========================================================

        if impresora_original:
            try:
                win32print.SetDefaultPrinter(impresora_original)
            except Exception:
                pass

def convertir_orden_excel_a_pdf(ruta_excel: str | Path) -> Path:
    """
    Convierte un Excel a PDF en la misma carpeta.

    Para cada hoja visible:
    - Imprime desde la columna A hasta la J.
    - Detecta la última fila con contenido dentro de A:J.
    - Añade una fila adicional.
    - Elimina los saltos de línea internos.
    - Ajusta automáticamente el ancho de las columnas.
    - Ajusta automáticamente el alto de las filas.
    - Ajusta el documento a una página de ancho.

    Requisitos:
        pip install pywin32
        Microsoft Excel instalado en Windows.
    """

    ruta_excel = Path(ruta_excel).expanduser().resolve()

    extensiones_validas = {
        ".xlsx",
        ".xlsm",
        ".xls",
        ".xlsb",
    }

    if ruta_excel.suffix.lower() not in extensiones_validas:
        raise ValueError(
            f"Formato no compatible: {ruta_excel.suffix}. "
            f"Formatos admitidos: "
            f"{', '.join(sorted(extensiones_validas))}"
        )

    if not ruta_excel.is_file():
        raise FileNotFoundError(
            f"No se encuentra el archivo: {ruta_excel}"
        )

    ruta_pdf = ruta_excel.with_suffix(".pdf")

    # Constantes de Microsoft Excel
    XL_TYPE_PDF = 0
    XL_QUALITY_STANDARD = 0

    XL_FORMULAS = -4123
    XL_PART = 2
    XL_BY_ROWS = 1
    XL_PREVIOUS = 2

    XL_SHEET_VISIBLE = -1

    XL_LANDSCAPE = 2
    XL_PAPER_A4 = 9

    excel = None
    libro = None

    pythoncom.CoInitialize()

    try:
        excel = win32.DispatchEx("Excel.Application")

        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        libro = excel.Workbooks.Open(
            str(ruta_excel),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
        )

        hojas_visibles = 0

        for indice in range(
            1,
            libro.Worksheets.Count + 1,
        ):
            hoja = libro.Worksheets(indice)

            if hoja.Visible != XL_SHEET_VISIBLE:
                continue

            hojas_visibles += 1

            rango_busqueda = hoja.Range("A:J")

            # Busca la última celda con contenido en A:J
            ultima_celda = rango_busqueda.Find(
                What="*",
                After=rango_busqueda.Cells(1, 1),
                LookIn=XL_FORMULAS,
                LookAt=XL_PART,
                SearchOrder=XL_BY_ROWS,
                SearchDirection=XL_PREVIOUS,
                MatchCase=False,
                SearchFormat=False,
            )

            ultima_fila_rellena = (
                int(ultima_celda.Row)
                if ultima_celda
                else 1
            )

            ultima_fila_impresion = (
                ultima_fila_rellena + 1
            )

            rango_impresion = hoja.Range(
                hoja.Cells(1, 1),
                hoja.Cells(
                    ultima_fila_impresion,
                    10,  # Columna J
                ),
            )

            # -------------------------------------------------
            # ELIMINAR SALTOS DE LÍNEA MANUALES
            # -------------------------------------------------

            for numero_fila in range(
                1,
                ultima_fila_impresion + 1,
            ):
                for numero_columna in range(1, 11):

                    celda = hoja.Cells(
                        numero_fila,
                        numero_columna,
                    )

                    valor = celda.Value

                    if isinstance(valor, str):
                        valor_sin_saltos = (
                            valor
                            .replace("\r\n", " ")
                            .replace("\n", " ")
                            .replace("\r", " ")
                        )

                        # Eliminar espacios repetidos
                        valor_sin_saltos = " ".join(
                            valor_sin_saltos.split()
                        )

                        if valor_sin_saltos != valor:
                            celda.Value = valor_sin_saltos

            # -------------------------------------------------
            # AJUSTAR CELDAS
            # -------------------------------------------------

            # Impedir que Excel divida el texto en varias líneas
            rango_impresion.WrapText = False

            # Ajustar el texto dentro de la celda si todavía
            # no entra, especialmente en celdas combinadas
            rango_impresion.ShrinkToFit = True

            # Ajustar automáticamente cada columna
            for numero_columna in range(1, 11):

                rango_columna = hoja.Range(
                    hoja.Cells(1, numero_columna),
                    hoja.Cells(
                        ultima_fila_impresion,
                        numero_columna,
                    ),
                )

                try:
                    rango_columna.Columns.AutoFit()
                except Exception:
                    # AutoFit puede fallar en columnas que
                    # contienen determinadas celdas combinadas
                    pass

                # Añadir un pequeño margen para evitar que
                # el texto quede justo al convertirlo a PDF
                columna_completa = hoja.Columns(
                    numero_columna
                )

                try:
                    ancho_actual = float(
                        columna_completa.ColumnWidth
                    )

                    columna_completa.ColumnWidth = min(
                        ancho_actual + 1.5,
                        255,
                    )

                except (TypeError, ValueError):
                    pass

            # Ajustar automáticamente el alto de las filas
            try:
                rango_impresion.Rows.AutoFit()
            except Exception:
                # Puede fallar si existen celdas combinadas
                pass

            # -------------------------------------------------
            # CONFIGURACIÓN DE IMPRESIÓN
            # -------------------------------------------------

            configuracion = hoja.PageSetup

            configuracion.PrintArea = (
                f"$A$1:$J${ultima_fila_impresion}"
            )

            # A4 horizontal
            configuracion.Orientation = XL_LANDSCAPE
            # configuracion.PaperSize = XL_PAPER_A4

            # Una página de ancho y las páginas de alto
            # que sean necesarias
            configuracion.Zoom = False
            configuracion.FitToPagesWide = 1
            configuracion.FitToPagesTall = False

            # Márgenes reducidos
            configuracion.LeftMargin = (
                excel.CentimetersToPoints(0.4)
            )

            configuracion.RightMargin = (
                excel.CentimetersToPoints(0.4)
            )

            configuracion.TopMargin = (
                excel.CentimetersToPoints(0.6)
            )

            configuracion.BottomMargin = (
                excel.CentimetersToPoints(0.6)
            )

            configuracion.HeaderMargin = (
                excel.CentimetersToPoints(0.2)
            )

            configuracion.FooterMargin = (
                excel.CentimetersToPoints(0.2)
            )

            configuracion.CenterHorizontally = True

        if hojas_visibles == 0:
            raise ValueError(
                "El libro no contiene ninguna hoja visible."
            )

        # Eliminar el PDF anterior, si ya existe
        if ruta_pdf.exists():
            ruta_pdf.unlink()

        libro.ExportAsFixedFormat(
            Type=XL_TYPE_PDF,
            Filename=str(ruta_pdf),
            Quality=XL_QUALITY_STANDARD,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False,
        )

        return ruta_pdf

    finally:
        if libro is not None:
            libro.Close(SaveChanges=False)

        if excel is not None:
            excel.Quit()

        pythoncom.CoUninitialize()

def normalizar_texto(valor):
    """
    Normaliza un texto para poder comparar cabeceras aunque
    tengan tildes, minúsculas o espacios adicionales.

    Ejemplo:
        "Orden de producción " -> "ORDEN DE PRODUCCION"
    """

    if pd.isna(valor):
        return ""

    texto = str(valor).strip().upper()

    texto = unicodedata.normalize(
        "NFKD",
        texto
    ).encode(
        "ascii",
        "ignore"
    ).decode("ascii")

    texto = re.sub(r"\s+", " ", texto)

    return texto

def ordenar_ordenes_produccion(ordenes_produccion_excel):
    """
    Ordena las órdenes de producción de menor a mayor
    usando el número inicial del nombre del archivo.

    Acepta rutas Path o cadenas de texto.

    Ejemplo:
        964 25_12_30.xlsx
        486 23_04_13.xlsx
        1002 26_01_10.xlsx
    """

    def obtener_numero_orden(archivo):
        nombre = Path(archivo).stem.strip()

        coincidencia = re.match(
            r"^(\d{3,4})",
            nombre
        )

        if coincidencia is None:
            return float("inf")

        return int(coincidencia.group(1))

    return sorted(
        ordenes_produccion_excel,
        key=obtener_numero_orden
    )

def detectar_fila_cabecera(df_bruto, max_filas_busqueda=50):
    """
    Busca la fila que contiene las cabeceras mínimas necesarias.

    No presupone que la cabecera esté en la fila 4.
    """

    cabeceras_necesarias = {
        "ORDEN DE PRODUCCION",
        "NS INIT",
        "NS FIN"
    }

    filas_encontradas = []

    limite = min(max_filas_busqueda,len(df_bruto))

    for indice in range(limite):

        valores_fila = {normalizar_texto(valor) for valor in df_bruto.iloc[indice].tolist()}

        if cabeceras_necesarias.issubset(valores_fila):
            filas_encontradas.append(indice)

    if not filas_encontradas:
        raise ValueError(
            "No se ha encontrado una cabecera válida. "
            "La fila debe contener 'ORDEN DE PRODUCCION', "
            "'NS INIT' y 'NS FIN'."
        )

    if len(filas_encontradas) > 1:
        raise ValueError(
            "Se han encontrado varias posibles filas de cabecera: "
            f"{[indice + 1 for indice in filas_encontradas]}"
        )

    return filas_encontradas[0]

def normalizar_nombres_archivos(directorio):
    """
    Normaliza los nombres de los archivos de una carpeta:

    - Elimina espacios al principio y al final.
    - Elimina espacios antes de la extensión.
    - Convierte varios espacios consecutivos en uno.
    - Elimina espacios alrededor de los guiones bajos.
    - Cambia - por _ en los nombres de los archivos

    Ejemplo:
        ' 555  24_02_14 .xlsx'
        pasa a:
        '555 24_02_14.xlsx'
    """

    directorio = Path(directorio)

    for archivo in directorio.iterdir():

        if not archivo.is_file():
            continue

        nombre_sin_extension = archivo.stem

        # Eliminar espacios iniciales y finales
        nombre_normalizado = nombre_sin_extension.strip()

        # Convertir varios espacios consecutivos en uno
        nombre_normalizado = re.sub(
            r"\s+",
            " ",
            nombre_normalizado
        )

        # Eliminar espacios alrededor de los guiones bajos
        nombre_normalizado = re.sub(
            r"\s*_\s*",
            "_",
            nombre_normalizado
        )

        nombre_normalizado = re.sub(
            r"[-‐-‒–—]",
            "_",
            nombre_normalizado
        )
        
        nuevo_nombre = nombre_normalizado + archivo.suffix
        nueva_ruta = archivo.with_name(nuevo_nombre)

        if nueva_ruta == archivo:
            continue

        if nueva_ruta.exists():
            print(
                f"No se puede renombrar '{archivo.name}': "
                f"ya existe '{nueva_ruta.name}'."
            )
            continue

        archivo.rename(nueva_ruta)

        print(
            f"Renombrado: '{archivo.name}' "
            f"→ '{nueva_ruta.name}'"
        )

def leer_excel_PNTs():
    """
    Carga el Excel de PNT y devuelve un DataFrame con los datos.
    """

    df_pnt = pd.read_excel(RUTA_PNT_EXCEL).ffill(axis=0)

    # Normalizamos los nombres de las columnas
    df_pnt.columns = [
        normalizar_texto(columna)
        for columna in df_pnt.columns
    ]
    
    columna = "No REGISTRO REGISTRATION NO."
    df_pnt[columna] = (
        df_pnt[columna]
        .astype(str)
        .str.extract(r"(\d{4})", expand=False)
    )
    
    columna = "REF / REF"
    df_pnt[columna] = (
        df_pnt[columna]
        .astype("string")
        .str.strip()
        .str.upper()
    )

    # Eliminar filas cuyo campo LOTE/LOT contenga devolución o reparación
    df_pnt = df_pnt[
        ~df_pnt["LOTE / LOT"].str.contains(
            r"devoluci[oó]n|reparaci[oó]n",
            case=False,  # Ignora mayúsculas y minúsculas
            na=False,    # Los valores NaN no se consideran coincidencias
            regex=True
        )
    ].copy()
    return df_pnt

def leer_configuracion():
    """
    Lee un archivo Excel y convierte cada columna en una lista.

    Devuelve
    --------
    dict
        Diccionario con el nombre de cada columna y su lista de valores.
    """

    df = pd.read_excel(RUTA_CONFIG)  # Leer el Excel
    listas_columnas = {} # Diccionario donde se almacenarán las listas

    for columna in df.columns:

        # Eliminar valores vacíos y convertir la columna en una lista
        lista = df[columna].dropna().tolist()

        # Guardar la lista utilizando como clave el nombre de la columna
        listas_columnas[columna] = lista

    return listas_columnas

def normalizar_numero_pnt(valor):
    """
    Evita que pandas devuelva números como 4485.0.
    """

    if pd.isna(valor):
        return ""

    try:
        return str(int(float(valor)))
    except (TypeError, ValueError):
        return str(valor).strip()

def normalizar_numero_lote(valor):
    """
    Obtiene el número final de un lote y lo convierte en entero.

    Ejemplos:
        14          -> 14
        "14"        -> 14
        "000014"    -> 14
        "25_000014" -> 14
        "EPB000014" -> 14

    Si no se puede interpretar, devuelve None.
    """

    if valor is None:
        return None

    texto = str(valor).strip()

    coincidencia = re.search(
        r"(\d+)\s*$",
        texto
    )

    if coincidencia is None:
        return None

    return int(coincidencia.group(1))

def crear_mascara_articulo(df_pnt, columna_referencia, nombre_articulo):
    """
    Devuelve una máscara booleana indicando qué filas contienen
    exactamente el artículo buscado.
    Las referencias pueden estar separadas por:
    - Saltos de línea
    - Comas
    - Punto y coma
    - Barra vertical
    """
    # df_pnt.to_excel(
    #     Path.home() / "Desktop" / f"PNTs filtrado.xlsx",
    #     index=False)
    # time.sleep(50)
    nombre_articulo_normalizado = (
        str(nombre_articulo)
        .replace("\xa0", " ")
        .strip()
        .upper()
    )
    def contiene_articulo(valor_referencia):
        if pd.isna(valor_referencia):
            return False
        referencias = [
            referencia
            .replace("\xa0", " ")
            .strip()
            .upper()
            for referencia in re.split(
                r"[\n,;|]+",
                str(valor_referencia)
            )
            if referencia.strip()
        ]
        return nombre_articulo_normalizado in referencias
    mascara_articulo = (
        df_pnt[columna_referencia]
        .apply(contiene_articulo)
    )
    return mascara_articulo

def interpretar_celda_lotes(valor_celda, max_elementos_rango=20000):
    """
    Interpreta una celda de la columna LOTE / LOT.

    Siempre devuelve una lista:

    - Celda vacía o error:
        []

    - Valor único:
        ["EPB220911"]

    - Varios valores:
        ["EP22479", "EP22480", "EP22481"]

    - Rango:
        ["EP1250056", "EP1250057", ..., "EP1250070"]

    - Rangos y valores mezclados:
        Devuelve todos los valores en una única lista.
    """

    try:
        # =====================================================
        # CASO 1: CELDA VACÍA
        # =====================================================

        if pd.isna(valor_celda):
            return []

        texto = str(valor_celda)

        # Normalizar espacios y diferentes tipos de guiones.
        texto = (
            texto
            .replace("\xa0", " ")
            .replace("–", "-")
            .replace("—", "-")
            .strip()
            .strip('"')
        )

        if not texto:
            return []

        # Corregir espacios alrededor de guiones bajos.
        #
        # "BSPDEPBV01 _000049"
        # se convierte en:
        # "BSPDEPBV01_000049"
        texto = re.sub(
            r"\s*_\s*",
            "_",
            texto
        )

        # Eliminar cantidades situadas delante de los lotes.
        #
        # "250 unid BHAREPBV01_000081-BHAREPBV01_000083"
        # pasa a:
        # "BHAREPBV01_000081-BHAREPBV01_000083"
        texto = re.sub(
            r"\b\d+\s*"
            r"(?:"
            r"unid(?:ades)?|"
            r"uds?|"
            r"uni|"
            r"und"
            r")"
            r"\b\s*:?",
            " ",
            texto,
            flags=re.IGNORECASE
        )

        # =====================================================
        # FUNCIÓN AUXILIAR PARA SEPARAR PREFIJO Y NÚMERO
        # =====================================================

        def separar_serie(serie):
            """
            Separa el prefijo y la parte numérica final.

            Ejemplo:
                EPB1250056 -> ("EPB", "1250056")
                25_000056  -> ("25_", "000056")
            """

            coincidencia = re.fullmatch(
                r"(?P<prefijo>.*?)(?P<numero>\d+)",
                serie
            )

            if coincidencia is None:
                return None

            return (
                coincidencia.group("prefijo"),
                coincidencia.group("numero")
            )

        # =====================================================
        # FUNCIÓN AUXILIAR PARA GENERAR UN RANGO
        # =====================================================

        def generar_rango(serie_inicial, serie_final):
            """
            Genera los valores comprendidos entre dos series.

            Si el rango no es válido, devuelve una lista vacía.
            """

            datos_iniciales = separar_serie(serie_inicial)
            datos_finales = separar_serie(serie_final)

            if datos_iniciales is None or datos_finales is None:
                return []

            prefijo_inicial, numero_inicial_texto = datos_iniciales
            prefijo_final, numero_final_texto = datos_finales

            # Los prefijos deben coincidir.
            if prefijo_inicial.upper() != prefijo_final.upper():
                return []

            numero_inicial = int(numero_inicial_texto)
            numero_final = int(numero_final_texto)

            # El valor final no puede ser menor que el inicial.
            if numero_final < numero_inicial:
                return []

            cantidad_elementos = (
                numero_final - numero_inicial + 1
            )

            # Evitar crear listas enormes por errores en los datos.
            if cantidad_elementos > max_elementos_rango:
                return []

            longitud_numero = max(
                len(numero_inicial_texto),
                len(numero_final_texto)
            )

            return [
                (
                    f"{prefijo_inicial}"
                    f"{numero:0{longitud_numero}d}"
                )
                for numero in range(
                    numero_inicial,
                    numero_final + 1
                )
            ]

        # =====================================================
        # DIVIDIR LA CELDA EN FRAGMENTOS
        # =====================================================

        # Divide por:
        # - comas
        # - punto y coma
        # - barra vertical
        # - saltos de línea
        fragmentos = [
            fragmento.strip(" .:")
            for fragmento in re.split(
                r"[\n,;|]+",
                texto
            )
            if fragmento.strip(" .:")
        ]

        valores = []

        # Patrón general de una serie.
        patron_serie = (
            r"[A-Za-z0-9+_/().]*\d+"
        )

        # Patrón de rango completo.
        patron_rango = re.compile(
            rf"^\s*"
            rf"(?P<inicial>{patron_serie})"
            rf"\s*-\s*"
            rf"(?P<final>{patron_serie})"
            rf"\s*$",
            flags=re.IGNORECASE
        )

        # Patrón de valor único.
        patron_valor_unico = re.compile(
            rf"^\s*({patron_serie})\s*$",
            flags=re.IGNORECASE
        )

        # =====================================================
        # ANALIZAR CADA FRAGMENTO
        # =====================================================

        for fragmento in fragmentos:

            # -----------------------------------------------
            # CASO 2: RANGO
            # -----------------------------------------------

            coincidencia_rango = patron_rango.fullmatch(
                fragmento
            )

            if coincidencia_rango is not None:

                serie_inicial = coincidencia_rango.group(
                    "inicial"
                )

                serie_final = coincidencia_rango.group(
                    "final"
                )

                valores_rango = generar_rango(
                    serie_inicial,
                    serie_final
                )

                # Si el rango es incorrecto, se ignora.
                if valores_rango:
                    valores.extend(valores_rango)

                continue

            # -----------------------------------------------
            # CASO 3: VALOR ÚNICO
            # -----------------------------------------------

            coincidencia_unica = patron_valor_unico.fullmatch(
                fragmento
            )

            if coincidencia_unica is not None:
                valores.append(
                    coincidencia_unica.group(1)
                )
                continue

            # -----------------------------------------------
            # CASO 4: VARIOS RANGOS O VALORES EN EL FRAGMENTO
            # -----------------------------------------------

            patron_elementos = re.compile(
                rf"(?P<inicial>{patron_serie})"
                rf"\s*-\s*"
                rf"(?P<final>{patron_serie})"
                rf"|"
                rf"(?P<unico>{patron_serie})",
                flags=re.IGNORECASE
            )

            for coincidencia in patron_elementos.finditer(
                fragmento
            ):

                serie_inicial = coincidencia.group(
                    "inicial"
                )

                serie_final = coincidencia.group(
                    "final"
                )

                valor_unico = coincidencia.group(
                    "unico"
                )

                if serie_inicial and serie_final:

                    valores_rango = generar_rango(
                        serie_inicial,
                        serie_final
                    )

                    if valores_rango:
                        valores.extend(valores_rango)

                elif valor_unico:
                    valores.append(valor_unico)

        # =====================================================
        # ELIMINAR DUPLICADOS
        # =====================================================

        resultado = []
        valores_vistos = set()

        for valor in valores:

            valor_limpio = (
                str(valor)
                .strip()
                .rstrip(".")
            )

            clave = valor_limpio.upper()

            if (
                valor_limpio
                and clave not in valores_vistos
            ):
                valores_vistos.add(clave)
                resultado.append(valor_limpio)

        return resultado

    except Exception:
        # Ante cualquier error inesperado, devolver lista vacía.
        return []

def buscar_pnt_articulo(df_pnt,componente,lote_numero_buscado):
    """
    Devuelve una tupla (numero_pnt, numero_albaran) asociada a un lote o numero de un componente/articulo.
    Busca el PNT asociado a un lote o numero de un componente/articulo.
    Permite encontrar el lote cuando en el registro PNT aparece:

    - Como lote individual:
        PINZBEPBV01_000104

    - Dentro de un intervalo:
        PINZBEPBV01_000102-PINZBEPBV01_000105

    - Con espacios:
        BHABEPBV01_000114 - BHABEPBV01_000117

    - Con un separador incorrecto:
        BHABEPBV01_000083_BHABEPBV01_000088

    La primera columna de df_pnt se considera siempre
    la columna que contiene el número de PNT.
    """

    
    columna_pnt = df_pnt.columns[0]
    columna_lote = "LOTE / LOT"
    columna_referencia = "REF / REF"
    columna_albaran = "ALBARAN / ALBARAN"
    
    if columna_lote not in df_pnt.columns:
        raise KeyError(
            f"No existe la columna '{columna_lote}' en df_pnt."
        )

    if columna_referencia not in df_pnt.columns:
        raise KeyError(
            f"No existe la columna '{columna_referencia}' en df_pnt."
        )

    componente_normalizado = str(componente).strip().upper()
    lote_buscado_texto = str(lote_numero_buscado).strip().upper()

    if lote_buscado_texto == "" or lote_buscado_texto == "NAN":
        lista_errores.append(f"{componente} con {lote_numero_buscado} no se ha encontrado en el PNT de calidad")
        return "",""

    # El argumento puede llegar como: BHABEPBV01_000114 o simplemente como: 000114
    coincidencia_lote = re.search(
        rf"(?:{re.escape(componente_normalizado)}\s*_\s*)?"
        r"(\d+)\s*$",
        lote_buscado_texto,
        flags=re.IGNORECASE
    )

    if coincidencia_lote is None:
        lista_errores.append(f"{componente} con {lote_numero_buscado} no se ha encontrado en el PNT de calidad")
        return "",""

    numero_lote_buscado = int(
        coincidencia_lote.group(1)
    )

    mascara_articulo = crear_mascara_articulo(
        df_pnt=df_pnt,
        columna_referencia="REF / REF",
        nombre_articulo=componente_normalizado
    )

    df_pnt_articulo = df_pnt.loc[mascara_articulo]
    # Comprobamos si existe el articulo en el dataframe
    if df_pnt_articulo.empty:
        print(f"El articulo {componente_normalizado} no se encuentra en el PNT")
        lista_errores.append(f"El articulo {componente_normalizado} no se encuentra en el PNT")
        # time.sleep(1)
        return ("","")
    
    # print("Valor que buscamos: ", numero_lote_buscado)
    for _, fila in df_pnt_articulo.iterrows():

        valor_celda = fila["LOTE / LOT"]
        lista_lotes = interpretar_celda_lotes(valor_celda) # Devuelve el lote a pelo 25_0032 o PHB_000235
        lista_lotes_normalizada = list(map(normalizar_numero_lote, lista_lotes))

        # print("Lista de lotes:",lista_lotes)

        if numero_lote_buscado in lista_lotes_normalizada:
              
            albaran = fila[columna_albaran]
            pnt = fila.iloc[0]

            return (
                normalizar_numero_pnt(pnt),
                "" if pd.isna(albaran)
                else str(albaran).strip()
            )
    
    # --------------- BLOQUE PARA DEPURAR ERRORES RELACIONADOS CON LOTES CONCRETOS -------------------------------------------
    # if(numero_lote_buscado == '105250554' or numero_lote_buscado==105250554 ):
    #     for _, fila in df_pnt_articulo.iterrows():
    #         # Para depurar solo hay prints
    #             print(
    #                 "\nLOCALIZADO\n",
    #                 "\n\nValor que buscamos:\n", numero_lote_buscado,
    #                 f"\nValor de la celda en la que se lee el rango de lotes dentro del pnt:",
    #                 fila["LOTE / LOT"],
    #                 "\n\n\n"
    #             )
    #             time.sleep(0.1)
    #         # print("Lista de lotes:\n",fila["LOTE / LOT"])
    #     time.sleep(50)

    lista_errores.append(f"No se ha encontrado {componente} con lote {numero_lote_buscado} en PNT  ")

    return ("","")

def buscar_pnt_orden(df_pnt, nombre_orden, nombre_articulo):
    """
    Busca el número de PNT correspondiente a una orden de producción y su articulo.

    Para considerar válida una fila deben coincidir:

    1. El número de orden en la columna 'ALBARAN / ALBARAN'.
    2. El artículo en la columna 'REF / REF'.

    Ejemplo:
        nombre_orden = "861 25_08_18.xlsx"
        nombre_articulo = "PHB1V01"
    """
    
    columna_albaran = "ALBARAN / ALBARAN"   # Columna que contiene tanto albaranes como ordenes de produccion (usada para la busqueda)
    columna_referencia = "REF / REF"        # Columnas que contiene el articulo (Usada para filtrar)
    columna_pnt = df_pnt.columns[0]         # Se considera que el número de PNT está en la primera columna

    # Comprobar que existen las columnas necesarias
    columnas_necesarias = [columna_albaran,columna_referencia]
    for columna in columnas_necesarias:
        if columna not in df_pnt.columns:
            raise KeyError(
                f"No existe la columna '{columna}' en df_pnt."
            )

    # ---------------------------------------------------------
    # Obtener el número de la orden
    # ---------------------------------------------------------

    # Eliminar la extensión del archivo
    # "861 25_08_18.xlsx" -> "861 25_08_18"
    
    # Extraer el número situado al principio
    # "861 25_08_18" -> "861"
    numero_orden = re.match(
        r"\s*(\d+)",
        nombre_orden
    )

    if numero_orden is None:
        mensaje_error = (
            f"No se puede obtener el número de orden de "
            f"'{nombre_orden}'."
        )
        lista_errores.append(mensaje_error)
        raise ValueError(mensaje_error)

    numero_orden = numero_orden.group(1)


    # ---------------------------------------------------------
    # Filtrar primero por artículo
    # ---------------------------------------------------------

    mascara_articulo = crear_mascara_articulo(
        df_pnt=df_pnt,
        columna_referencia="REF / REF",
        nombre_articulo=nombre_articulo
    )

    df_pnt_articulo = df_pnt.loc[mascara_articulo]

    # df_pnt_articulo.to_excel(
    #     Path.home() / "Desktop" / f"PNTs filtrado.xlsx",
    #     index=False)
    # time.sleep(50)
    
    # Si no hay ninguna fila para ese artículo, registrar el error
    if df_pnt_articulo.empty:
        mensaje_error = (
            f"No se encuentra ningún PNT para el artículo "
            f"{nombre_articulo}."
            f"En la orden {nombre_orden}."
        )
        lista_errores.append(mensaje_error)
        

    def busca_pnt_df(df_buscar):
    
        # -----------------------------------------------------------------
        # Buscar la orden en df filtrado por el articulo la orden deseada
        # -----------------------------------------------------------------

        # 0* permite que 861 coincida también con 0861 o 000861.
        #
        # (?<!\d) y (?!\d) evitan que 861 coincida dentro
        # de números como 1861 o 8610.
        patron_numero_orden = (
            rf"(?<!\d)0*{re.escape(numero_orden)}(?!\d)"
        )

        mascara_orden = (
            df_buscar[columna_albaran]
            .fillna("")
            .astype(str)
            .str.contains(
                patron_numero_orden,
                regex=True,
                na=False
            )
        )

        coincidencias = df_buscar.loc[mascara_orden]

        # ---------------------------------------------------------
        # Devolver el número de PNT
        # ---------------------------------------------------------

        if not coincidencias.empty:
            return normalizar_numero_pnt(
                coincidencias.iloc[0][columna_pnt]
            )

        return ""
    
    pnt = busca_pnt_df(df_pnt_articulo) # Primero la buscamos en el df filtrado por el articulo
    
    if pnt == "":
        # Si no se encuentra el pnt habiendo filtrado el articulo, lo buscamos en todo el df de pnts
        mensaje_error = (
            f"No se encuentra el PNT de la orden "
            f"{numero_orden} para el artículo "
            f"{nombre_articulo}. dentro de los articulos {nombre_articulo}"
        )
        lista_errores.append(mensaje_error)

        pnt = busca_pnt_df(df_pnt)

        if pnt == "":
            mensaje_error = (
                f"No se encuentra el PNT de la orden "
                f"{numero_orden} en todo el pnt de control de calidad"
            )
            lista_errores.append(mensaje_error)
            return ""
        
    return pnt

def generar_numeros_serie(serie_inicial, serie_final, nombre_orden, nombre_articulo):
    """
    Genera todos los números de serie comprendidos entre
    una serie inicial y una serie final.

    Admite prefijos que contengan:
    - Letras
    - Números
    - Guiones bajos
    - Signos como +, x, /, etc.

    Ejemplos:
        EP1250056 - EP1250070
        BATEPBV01_000053 - BATEPBV01_000056
        22_001001 - 22_001312
        TM2x4EPBV01_000049 - TM2x4EPBV01_000064
        I+D_000008 - I+D_000011
        240001 - 240016
    """

    # ---------------------------------------------------------
    # 1. Normalizar los valores recibidos
    # ---------------------------------------------------------

    serie_inicial = str(serie_inicial).strip()
    serie_final = str(serie_final).strip()

    # Eliminar espacios introducidos accidentalmente
    # dentro de los números de serie.
    serie_inicial = re.sub(r"\s+", "", serie_inicial)
    serie_final = re.sub(r"\s+", "", serie_final)

    # ---------------------------------------------------------
    # 2. Separar prefijo y parte numérica final
    # ---------------------------------------------------------

    patron_serie = re.compile(
        r"^(?P<prefijo>.*?)(?P<numero>\d+)$"
    )

    coincidencia_inicial = patron_serie.fullmatch(
        serie_inicial
    )

    coincidencia_final = patron_serie.fullmatch(
        serie_final
    )

    if ( coincidencia_inicial is None or coincidencia_final is None):
        msg_error = (
            f"Los números de serie del articulo {nombre_articulo} no tienen un formato válido. "
            f"Orden del articulo {nombre_orden}"
            f"Valores recibidos: '{serie_inicial}' y "
            f"'{serie_final}'."
        )
        lista_errores.append(msg_error)
        raise ValueError(msg_error)
    # ---------------------------------------------------------
    # 3. Obtener prefijos y números
    # ---------------------------------------------------------

    prefijo_inicial = coincidencia_inicial.group(
        "prefijo"
    )

    numero_inicial_texto = coincidencia_inicial.group(
        "numero"
    )

    prefijo_final = coincidencia_final.group(
        "prefijo"
    )

    numero_final_texto = coincidencia_final.group(
        "numero"
    )

    # ---------------------------------------------------------
    # 4. Comprobar que los prefijos sean iguales
    # ---------------------------------------------------------

    if prefijo_inicial.upper() != prefijo_final.upper():
        msg_error = (
            f"Los números de serie del articulo {nombre_articulo} con "
            f"orden  {nombre_orden}-->"
            "La serie inicial y la serie final tienen "
            "prefijos diferentes: "
            f"'{prefijo_inicial}' y '{prefijo_final}'."
        )
        lista_errores.append(msg_error)
        raise ValueError(msg_error)


    # ---------------------------------------------------------
    # 5. Convertir la parte numérica a entero
    # ---------------------------------------------------------

    numero_inicial = int(numero_inicial_texto)
    numero_final = int(numero_final_texto)

    if numero_final < numero_inicial:
        msg_error = (
            f"Los números de serie del articulo {nombre_articulo} con "
            f"orden  {nombre_orden}-->"
            "La serie final es menor que la serie inicial: "
            f"{numero_inicial} > {numero_final}."
        )
        lista_errores.append(msg_error)
        return msg_error

    # ---------------------------------------------------------
    # 6. Determinar la cantidad de dígitos
    # ---------------------------------------------------------

    longitud_numero = max(
        len(numero_inicial_texto),
        len(numero_final_texto)
    )

    # ---------------------------------------------------------
    # 7. Generar todas las series
    # ---------------------------------------------------------
    
    numeros_serie = [
        f"{prefijo_inicial}{numero:0{longitud_numero}d}"
        for numero in range(
            numero_inicial,
            numero_final + 1
        )
    ]

    if(len(numeros_serie)>10000):
        msg_error = (
            f"La longitud de los numeros de serie es excesiva y supera los "
            f"10000, asegurate de que no hay ningun erro en la orden {nombre_orden} del articulo {nombre_articulo}."
        )
        lista_errores.append(msg_error)
        raise ValueError(msg_error)

    return numeros_serie

def distribuir_componente(numeros_serie,df_componente):
    """
    POR TESTEAR:  Distribuye el consumo total de un componente entre todos los
    números de serie.
    Distribuye el consumo total de un componente entre todos los
    números de serie.

    Los lotes se consumen en el orden en que aparecen en el Excel.
    """

    numero_equipos = len(numeros_serie)

    if numero_equipos == 0:
        raise ValueError("No existen números de serie.")

    df_componente = df_componente.copy()

    df_componente["UNIDADES_REALES"] = pd.to_numeric(
        df_componente["UNIDADES_REALES"],
        errors="coerce"
    ).fillna(0)

    df_componente["SERIE_LOTE"] = (
        df_componente["SERIE_LOTE"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Agrupamos los lotes manteniendo el orden de aparición
    consumos_por_lote = (
        df_componente
        .groupby(
            "SERIE_LOTE",
            sort=False
        )["UNIDADES_REALES"]
        .sum()
    )

    total_consumido = consumos_por_lote.sum()

    consumo_por_equipo = total_consumido / numero_equipos

    lotes = [
        {
            "LOTE": lote,
            "UNIDADES_DISPONIBLES": float(unidades)
        }
        for lote, unidades in consumos_por_lote.items()
    ]

    resultado = {}

    indice_lote = 0
    tolerancia = 1e-9

    for numero_serie in numeros_serie:

        unidades_pendientes = float(consumo_por_equipo)
        consumos_serie = []

        while unidades_pendientes > tolerancia:

            # Avanzamos si el lote actual se ha agotado
            while (
                indice_lote < len(lotes)
                and lotes[indice_lote]["UNIDADES_DISPONIBLES"]
                <= tolerancia
            ):
                indice_lote += 1

            if indice_lote >= len(lotes):

                raise ValueError(
                    f"{numeros_serie},{df_componente}",
                    f"No hay suficientes unidades en el lote {indice_lote} para completar "
                    f"la serie {numero_serie}."
                )

            lote_actual = lotes[indice_lote]

            unidades_asignadas = min(
                unidades_pendientes,
                lote_actual["UNIDADES_DISPONIBLES"]
            )

            consumos_serie.append({
                "LOTE": lote_actual["LOTE"],
                "UNIDADES": unidades_asignadas
            })

            lote_actual["UNIDADES_DISPONIBLES"] -= unidades_asignadas
            unidades_pendientes -= unidades_asignadas

        resultado[numero_serie] = consumos_serie

    return resultado

def crear_df_consumos_por_nserie(orden_produccion, nombre_articulo, df_produccion, numeros_serie, componentes):
    """
    Crea un DataFrame con una fila por número de serie o más si ese numero de serie requiere unidades de múltiples lotes.

    Para cada componente incluye:
    - Lote utilizado.
    - Unidades consumidas.
    """

    df = df_produccion.copy()

    df["CODIGO"] = (df["CODIGO"].fillna("").astype(str).str.strip())

    df["SERIE_LOTE"] = (df["SERIE_LOTE"].fillna("").astype(str).str.strip())

    df["UNIDADES_REALES"] = pd.to_numeric(df["UNIDADES_REALES"], errors="coerce").fillna(0)

    distribuciones = {}

    for componente in componentes:

        df_componente = df[df["CODIGO"] == componente].copy()

        if df_componente.empty:
            distribuciones[componente] = {numero_serie: []for numero_serie in numeros_serie}

        else:
            distribuciones[componente] = distribuir_componente(numeros_serie=numeros_serie,df_componente=df_componente)
    
    # Ahora creamos las filas del DataFrame
    filas_resultado = []

    for numero_serie in numeros_serie:

        # Calculamos cuántas filas necesita este número de serie.
        # Normalmente será una, pero será dos cuando algún
        # componente proceda de dos lotes.
        numero_filas_serie = 1

        for componente in componentes:

            consumos = distribuciones[
                componente
            ].get(numero_serie, [])

            numero_filas_serie = max(
                numero_filas_serie,
                len(consumos)
            )

        # Creamos una o varias filas repitiendo el número de serie
        for posicion_consumo in range(numero_filas_serie):

            if(nombre_articulo not in CONSOLAS):
                fila = {
                    "ORDEN DE PRODUCCION": orden_produccion,
                    "NUMERO_SERIE": numero_serie,
                    "UNIDADES": 1,
                    f"{nombre_articulo}_PNT": "",
                }
            else:
                fila = {
                    "ORDEN DE PRODUCCION": orden_produccion,
                    "NUMERO_SERIE": numero_serie,
                    "UNIDADES": 1,
                }

            
            for componente in componentes:
                
                consumos = distribuciones[
                    componente
                ].get(numero_serie, [])

                if posicion_consumo < len(consumos):

                    consumo = consumos[posicion_consumo]

                    fila[f"{componente}_LOTE"] = (
                        consumo["LOTE"]
                    )

                    fila[f"{componente}_UNID"] = (
                        consumo["UNIDADES"]
                    )
                    fila[f"{componente}_ALBARAN"] = ""
                    
                    if(componente not in CONSOLAS):
                        fila[f"{componente}_PNT"] = ""
                else:
                    # Este componente no necesita otra fila.
                    # Se deja vacío para no duplicar el consumo.
                    fila[f"{componente}_LOTE"] = ""
                    fila[f"{componente}_UNID"] = ""
                    fila[f"{componente}_ALBARAN"] = ""
                    if(componente not in CONSOLAS):
                        fila[f"{componente}_PNT"] = ""

            filas_resultado.append(fila)

    df_resultado = pd.DataFrame(filas_resultado)

    return df_resultado

def procesar_orden(ruta_excel):
    '''
    Lee una orden de producción con pandas.

    Devuelve:
    - Un DataFrame con el resumen de la orden de producción que se tiene que añadir al resumen de ordenes de producción.
    - Este DataFrame contiene una fila por número de serie y por cada componente se indica el PNT y el albaran que corresponda
    '''

    ruta_excel = Path(ruta_excel)

    # Leemos el Excel sin establecer ninguna fila como cabecera
    df_excel = pd.read_excel(
        ruta_excel,
        header=None
    )

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 1. LEER ARTÍCULO Y SERIES DE LA FILA 6
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    texto_orden = df_excel.iloc[5, 3] # Extraemos el texto que contien numeros de serie y nombre del articulo índice 5 en pandas,Columna D = índice 3

    texto_orden = str(texto_orden).strip()
    coincidencia = patron_lotes.search(texto_orden)

    if coincidencia is None:
        mensaje_error = (
            f"Hay algun error con el excel de la orden: {texto_orden}"
        )
        return None

    nombre_orden = os.path.splitext(ruta_excel.name)[0]
    # nombre_articulo = coincidencia.group("articulo").strip()
    nombre_articulo = ruta_excel.parent.name.strip()
    serie_inicial = coincidencia.group("serie_inicial").strip()
    serie_final = coincidencia.group("serie_final").strip()

    print(f"Procesnado Orden {nombre_orden} - Artículo: {nombre_articulo}, Serie Inicial: {serie_inicial}, Serie Final: {serie_final}")
    
    try:
        numeros_serie = generar_numeros_serie(serie_inicial, serie_final, nombre_orden, nombre_articulo)
    
    except:
        
        print(lista_errores[-1])
        time.sleep(2)
        return None
    # print(len(numeros_serie))
    # time.sleep(1)
    
    df_produccion = df_excel.iloc[
        9:,
        [0, 1, 4, 5, 7, 8]
    ].copy()

    df_produccion.columns = [
        "CODIGO",
        "DESCRIPCION",
        "SERIE_LOTE",
        "ALMACEN",
        "UNIDADES_PREVISTAS",
        "UNIDADES_REALES"
    ]

    # Eliminamos filas completamente vacías
    df_produccion = df_produccion.dropna(
        how="all"
    )

    # Eliminamos filas que no tengan código
    df_produccion = df_produccion.dropna(
        subset=["CODIGO"]
    )

    # Limpiamos espacios
    df_produccion["CODIGO"] = (
        df_produccion["CODIGO"]
        .astype(str)
        .str.strip()
    )

    df_produccion["DESCRIPCION"] = (
        df_produccion["DESCRIPCION"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # # Contamos cuántas veces aparece el artículo principal en la columna para saber el numero de agrupaciones de la produccion
    # agrupaciones_produccion = (df_produccion["CODIGO"] == nombre_articulo).sum()

    df_produccion["SERIE_LOTE"] = (
        df_produccion["SERIE_LOTE"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Marcamos las filas que corresponden al equipo fabricado
    df_produccion["ES_ARTICULO_PRINCIPAL"] = (
        df_produccion["CODIGO"] == nombre_articulo
    )

    df_produccion = df_produccion.reset_index(drop=True)

    componentes_articulo = (
        df_produccion.loc[
            ~df_produccion["ES_ARTICULO_PRINCIPAL"],
            "CODIGO"
        ]
        .dropna()
        .astype(str)
        .str.strip()
        .drop_duplicates()
        .tolist()
    )

    # Creamos un DataFrame con los consumos por número de serie que se añadira al resumen de ordenes de produccion
    componentes_articulo = sorted(componentes_articulo, key=lambda elemento: '_CON' not in elemento) # PONEMOS LA CONSOLA EN EL PRIMER ELEMENTO SIEMPRE
    
    
    df_consumos = crear_df_consumos_por_nserie(
        orden_produccion=nombre_orden,
        nombre_articulo=nombre_articulo,
        df_produccion=df_produccion,
        numeros_serie=numeros_serie,
        componentes=componentes_articulo
    )

    # Buscamos el PNT del artículo principal y lo añadimos al DataFrame

    if(nombre_articulo not in CONSOLAS):
        pnt_articulo = buscar_pnt_orden(df_pnt=DF_PNT, nombre_orden=nombre_orden, nombre_articulo=nombre_articulo)
        df_consumos[f"{nombre_articulo}_PNT"] = pnt_articulo

    # Buscamos el PNT y el albarán de cada componente y los añadimos al DataFrame de consumos
    # print(componentes_articulo)

    for componente in componentes_articulo:
        if componente not in CONSOLAS:
            columna_lote = f"{componente}_LOTE"

            if(componente not in CONSOLAS):
                columna_pnt = f"{componente}_PNT"

            columna_albaran = f"{componente}_ALBARAN"

            # Obtenemos solamente los lotes distintos del componente
            lotes_unicos = (df_consumos[columna_lote].dropna().astype(str).str.strip())
            lotes_unicos = lotes_unicos[lotes_unicos != ""].unique()

            # Buscamos el PNT y el albarán una sola vez por cada lote

            mapa_lote_datos = {
                lote: buscar_pnt_articulo(df_pnt=DF_PNT, componente=componente, lote_numero_buscado=lote) 
                for lote in lotes_unicos
            }

            # Separamos los PNT y los albaranes en dos diccionarios
            mapa_lote_pnt = {lote: datos[0] for lote, datos in mapa_lote_datos.items()}

            mapa_lote_albaran = {lote: datos[1] for lote, datos in mapa_lote_datos.items()}

            # Normalizamos la columna de lotes una sola vez
            serie_lotes = df_consumos[columna_lote].fillna("").astype(str).str.strip()

            # Asignamos el PNT
            if(componente not in CONSOLAS):
                df_consumos[columna_pnt] = (serie_lotes.map(mapa_lote_pnt).fillna(""))

            # Asignamos el albarán
            df_consumos[columna_albaran] = (serie_lotes.map(mapa_lote_albaran).fillna(""))
        
        else:
            consola_dispositivo = sorted(componentes_articulo, key=lambda elemento: '_CON' not in elemento)[0]
            ruta_excel_consola =  RUTA_IMD + consola_dispositivo +"/"+ f"{consola_dispositivo}_RESUMEN.xlsx" 
            df_consola = pd.read_excel(ruta_excel_consola).drop_duplicates(subset="NUMERO_SERIE", keep="first")


            df_consumos[f"{consola_dispositivo}_ALBARAN"] = df_consumos["NUMERO_SERIE"].map(
                df_consola.set_index("NUMERO_SERIE")["ORDEN DE PRODUCCION"]
            )

    # PARA LAS CONSOLAS ADEMAS AÑADIMOS OTRAS COLUMNAS ESPECIALES

    if(nombre_articulo  in DISPOSITIVOS):

        # PONEMOS A LA DERECHA DE LA OCLUMA DE LA CONSOLA LAS PCBs del dispositivo
        consola_dispositivo = sorted(componentes_articulo, key=lambda elemento: '_CON' not in elemento)[0]
        columna_referencia = f"{consola_dispositivo}_ALBARAN" # Gastamos de referencia la columna de la consola
        posicion = df_consumos.columns.get_loc(columna_referencia) + 1
        
        # OBTENEMOS EL DATAFRAME DE LA CONSOLA DEL EQUIPO
        ruta_excel_consola =  RUTA_IMD + consola_dispositivo +"/"+ f"{consola_dispositivo}_RESUMEN.xlsx"
        df_consola = pd.read_excel(ruta_excel_consola).drop_duplicates(subset="NUMERO_SERIE", keep="first")
        
        # AÑADIMOS EL NUMERO DE LA CONSOLA/COSOLAS AL DF DE CONSUMOS DEL DISPOSTIVO CORRESPONDIENTE
        
        # PROCESAMOS NUMERO DE LAS PCB:
        if(nombre_articulo in ["EPTEV02DEV01","EPTEV02DEV02"]):
            print("Añadiendo consolas")

            # Insertamos primero las dos columnas , la de PCP y la de PCBC
            df_consumos.insert(
                loc=posicion,
                column="PCBEPBP",
                value=None
            )
            df_consumos.insert(
                loc=posicion + 1 ,
                column="PCBEPBC",
                value=None
            )
            # Una vez insertadas las rellenamos PCBP con el valor que coincida en el excel de la consola
            df_consumos["PCBEPBP"] = df_consumos["NUMERO_SERIE"].map(
                df_consola.set_index("NUMERO_SERIE")[f"PCBEPBP_LOTE"]
            )
            # Una vez insertadas las rellenamos PCBC con el valor que coincida en el excel de la consola
            df_consumos["PCBEPBC"] = df_consumos["NUMERO_SERIE"].map(
                df_consola.set_index("NUMERO_SERIE")[f"PCBEPBC_LOTE"]
            ) 
        elif (nombre_articulo in ["EPTEV01DEV01","EPTEV01DEV02"]):# PARA EQUIPOS EPTEV01 SOLO HAY UNA CONSOLA

            df_consumos.insert(
                loc=posicion + 1,
                column="PCBEPV01",
                value=None
            )
            df_consumos["PCBEPV01"] = df_consumos["NUMERO_SERIE"].map(
                df_consola.set_index("NUMERO_SERIE")[f"PCBEPV01_LOTE"]
            )

        # PROCESAMOS CHECKLIST Y FECHA DE FABRICACION TODO
        # PROCESAMOS ALBARAN DE SALIDA
        # PROCESAMOS SI TIENE ETIQUETAS  (COLUMNA IDIOMAS ALBARAN Y REGDE LAS ETIQEUTAS)
        
    return df_consumos

def leer_excel_resumen(directorio_articulo):
    """
    Busca y lee el Excel que contiene la palabra 'resumen'
    dentro de la carpeta del artículo.
    Argumento: directorio de la carpeta donde se encuentra el excel resumen
    Devuelve : df_resumen , ordenes_produccion ya procesadas, ruta_resumen
    """

    directorio_articulo = Path(directorio_articulo)

    archivos_resumen = [
        archivo
        for archivo in directorio_articulo.iterdir()
        if archivo.is_file()
        and "resumen" in archivo.name.lower()
        and "formateado" not in archivo.name.lower()
        and archivo.suffix.lower() in (".xlsx", ".xls")
        and not archivo.name.startswith("~$")
    ]

    if not archivos_resumen:
        raise FileNotFoundError(
            f"No se ha encontrado ningún Excel resumen en "
            f"'{directorio_articulo}'."
        )

    if len(archivos_resumen) > 1:
        raise ValueError(
            "Se ha encontrado más de un Excel resumen: "
            f"{[archivo.name for archivo in archivos_resumen]}"
        )

    ruta_resumen = archivos_resumen[0]

    print(f"Leyendo Excel resumen: {ruta_resumen}")

    df_resumen = pd.read_excel(ruta_resumen)

    # Limpiamos los nombres de las columnas
    df_resumen.columns = (
        df_resumen.columns
        .astype(str)
        .str.strip()
    )

    # Eliminamos filas completamente vacías
    df_resumen = (
        df_resumen
        .dropna(how="all")
        .reset_index(drop=True)
    )

    # Las celdas combinadas de la orden aparecen como NaN.
    # Extendemos la orden hacia abajo.
    if "ORDEN DE PRODUCCION" in df_resumen.columns:
        df_resumen["ORDEN DE PRODUCCION"] = (
            df_resumen["ORDEN DE PRODUCCION"]
            .ffill()
        )
        
    ordenes_produccion = (
        df_resumen["ORDEN DE PRODUCCION"]
        .dropna()
        .astype(str)
        .str.strip()
        .drop_duplicates()
        .tolist()
    )

    return  df_resumen , ordenes_produccion, ruta_resumen

def obtener_ordenes_por_procesar(
    ordenes_produccion_excel,
    ordenes_procesadas
):
    """
    Devuelve las órdenes pendientes manteniendo exactamente
    el formato original de ordenes_produccion_excel.
    """

    procesadas_normalizadas = {
        str(orden).strip().replace("-", "_")
        for orden in ordenes_procesadas
    }

    ordenes_por_procesar = [
        orden_excel
        for orden_excel in ordenes_produccion_excel
        if (
            Path(orden_excel)
            .stem
            .strip()
            .replace("-", "_")
            not in procesadas_normalizadas
        )
    ]

    return ordenes_por_procesar

def obtener_requisitos_especiales(dispositivo, ns_ini, ns_final):
    """
    Devuelve un diccionario con los requisitos especiales de un artículo
    en función de su número de serie inicial y final.
    """

    ns_ini = str(ns_ini)[-4:]       # Solo ncesitamos los ultimos 4 digitos
    ns_final = str(ns_final)[-4:]   # Solo ncesitamos los ultimos 4 digitos
    def convertir_lista_en_rangos(ns_inicio,ns_final,lista_ns,requisito):
        """
        Filtra y agrupa números de serie usando únicamente
        los 4 últimos dígitos XXXX.

        Ejemplo:
            ns_inicio = "EPB1250442"  -> 0442
            ns_final  = "EPB1260600"  -> 0600

        lista_ns puede contener:
            EPB1250442
            250442
            0442
            etc.

        Devuelve:
            [
                {
                    "inicio": "0442",
                    "final": "0445",
                    "requisito": "INGLES",
                },
                ...
            ]
        """

        def obtener_4_digitos(valor):
            """
            Devuelve los 4 últimos dígitos del valor.
            """

            valor = str(valor).strip()

            # Por si pandas/Excel devuelve algo como 250442.0
            if valor.endswith(".0"):
                valor = valor[:-2]

            if len(valor) < 4:
                raise ValueError(
                    f"El valor debe contener al menos 4 caracteres: {valor}"
                )

            numero = valor[-4:]

            if not numero.isdigit():
                raise ValueError(
                    f"Los últimos 4 caracteres deben ser numéricos: {valor}"
                )

            return int(numero)


        # ========================================================
        # RANGO GENERAL
        # ========================================================

        numero_inicio = obtener_4_digitos(ns_inicio)
        numero_final = obtener_4_digitos(ns_final)

        if numero_inicio > numero_final:
            raise ValueError(
                f"El número inicial {numero_inicio:04d} "
                f"es mayor que el final {numero_final:04d}"
            )


        # ========================================================
        # FILTRAR LOS NS QUE ESTÁN DENTRO DEL RANGO
        # ========================================================

        numeros = []

        for ns in lista_ns:

            try:
                numero = obtener_4_digitos(ns)

            except ValueError:
                continue

            if numero_inicio <= numero <= numero_final:
                numeros.append(numero)


        # Eliminar duplicados y ordenar
        numeros = sorted(set(numeros))


        if not numeros:
            return []


        # ========================================================
        # CREAR RANGOS CONSECUTIVOS
        # ========================================================

        rangos = []

        inicio_rango = numeros[0]
        anterior = numeros[0]


        for numero in numeros[1:]:

            if numero == anterior + 1:

                anterior = numero

            else:

                rangos.append({
                    "inicio": f"{inicio_rango:04d}",
                    "final": f"{anterior:04d}",
                    "requisito": requisito,
                })

                inicio_rango = numero
                anterior = numero


        # Último rango
        rangos.append({
            "inicio": f"{inicio_rango:04d}",
            "final": f"{anterior:04d}",
            "requisito": requisito,
        })

        return rangos

    dispositivos_especiales = os.listdir(RUTA_ETIQUETAS)
    if dispositivo not in dispositivos_especiales:
        print(f"El dispositivo {dispositivo} no se encuentran las etiquetas impresas en otros idiomas.")
        return None

    idiomas_dispositivo  = os.listdir(RUTA_ETIQUETAS + dispositivo +'/')
    requisitos = []
    for idioma in idiomas_dispositivo:
        if "ES" in idioma or "desktop.ini" in idioma: # ESPAÑA NO TIENE REQUISITOS ESPECIALES, SOLO LOS DEMAS IDIOMAS
            continue
        if "EN" in idioma:
            txt_pais = "INGLES"
        if "FR" in idioma:
            txt_pais = "FRANCES"
        if "IT" in idioma:
            txt_pais = "ITALIANO"

        excel = obtener_unico_excel(RUTA_ETIQUETAS + dispositivo +'/'+ idioma +'/') # En caso de haber + d 1 excel peta
        df_etiquetas = pd.read_excel(excel)
        numeros_serie = [str(sn)[-4:]for sn in df_etiquetas["SN"].tolist()]
        # print(numeros_serie)
        print(ns_ini,ns_final)
        requisitos_idioma = convertir_lista_en_rangos(ns_ini, ns_final, numeros_serie, txt_pais)
        print(requisitos_idioma)
        if requisitos_idioma:
            requisitos.append(requisitos_idioma)

    lista_final = [
        diccionario
        for sublista in requisitos
        for diccionario in sublista
    ]

    # Como trabajamos unicamente con los ultimos 4 digitos necesitamos recuperar el numero de serie completo
    ruta_carpeta_articulo =  Path(RUTA_IMD + dispositivo +'/')
    df_resumen, _, _ = leer_excel_resumen(ruta_carpeta_articulo)
    # print(df_resumen)
    numeros_serie_completos = (df_resumen["NUMERO_SERIE"].dropna().astype(str).unique().tolist())

    def buscar_numero_serie_completo(ns_incompleto, lista_completos):

        # Nos quedamos con XXXX
        xxxx = str(ns_incompleto).strip()[-4:]

        coincidencias = []

        for ns in lista_completos:

            ns = str(ns).strip()

            # Por si Excel ha generado "EPB1260780.0"
            if ns.endswith(".0"):
                ns = ns[:-2]

            if ns[-4:] == xxxx:
                coincidencias.append(ns)

        if len(coincidencias) == 0:
            raise ValueError(
                f"No se ha encontrado el NS completo para {ns_incompleto} en el resumen de la orden de producción del artículo {dispositivo}."
            )

        if len(coincidencias) > 1:
            raise ValueError(
                f"Hay varios NS completos para {ns_incompleto}: "
                f"{coincidencias}"
            )

        return coincidencias[0]

    for item in lista_final:

        item["inicio"] = buscar_numero_serie_completo(item["inicio"],numeros_serie_completos,)

        item["final"] = buscar_numero_serie_completo(
            item["final"],
            numeros_serie_completos,
        )
            
    return lista_final

def filtrar_filas_entre_valores(df,columna_filtro,valor_inicio,valor_final):
    """
    Devuelve df de las filas comprendidas entre la primera aparición
    de valor_inicio y la primera aparición posterior de valor_final,
    incluyendo ambas filas.
    """

    # Buscar índice de inicio la posciion donde se cumple que se encuentra el valor inicial
    indices_inicio = df.index[df[columna_filtro] == valor_inicio].tolist()

    if not indices_inicio:
        raise ValueError(
            f"No se ha encontrado el valor inicial: {valor_inicio}"
        )

    indice_inicio = indices_inicio[0] # Nos quedamos con la primera aparicion

    # Buscar valor final a partir del inicio
    indices_final = df.index[(df[columna_filtro] == valor_final)& (df.index >= indice_inicio)].tolist()

    if not indices_final:
        raise ValueError(
            f"No se ha encontrado el valor final: {valor_final}"
        )

    indice_final = indices_final[-1] # Nos quedamos con la ultima aparicion

    # Filtrar incluyendo ambas filas
    return df.loc[indice_inicio:indice_final].copy()

def formatear_resumen_excel(ruta_excel, ruta_salida=None, hoja=None, device=None):
    """
    Formatea el Excel resumen agrupando visualmente la información.

    Operaciones realizadas
    ----------------------
    1. Busca la columna "ORDEN DE PRODUCCION".
    2. Combina las filas consecutivas que tengan la misma OP.
    3. Suma las UNIDADES correspondientes a cada OP.
    4. Combina las celdas de UNIDADES del mismo bloque y muestra
       únicamente el total.
    5. Combina las celdas consecutivas de EPTEV02DEV01_PNT
       cuando tengan el mismo valor.

    Las combinaciones se realizan únicamente sobre valores consecutivos.
    Si un mismo valor aparece nuevamente más adelante, separado por otro
    valor distinto, se considera un bloque diferente.

    Parámetros
    ----------
    ruta_excel : str | Path
        Excel de entrada.

    ruta_salida : str | Path, opcional
        Excel que se generará. Si no se indica, se crea un archivo
        con sufijo "_FORMATEADO".

    hoja : str, opcional
        Nombre de la hoja. Si no se indica se utiliza la hoja activa.

    Returns
    -------
    Path
        Ruta del Excel generado.
    """
    ruta_excel = Path(ruta_excel)
    # print(f"Formatenado excel {ruta_excel.name}")

    # ========================================================
    # RUTA DE SALIDA
    # ========================================================

    if ruta_salida is None:

        ruta_salida = ruta_excel.with_name(ruta_excel.stem+ "_FORMATEADO"+ ruta_excel.suffix)

    else:

        ruta_salida = Path(ruta_salida)

    # ========================================================
    # ABRIR EXCEL
    # ========================================================

    wb = load_workbook(ruta_excel)

    if hoja is None:
        ws = wb.active
    else:
        ws = wb[hoja]

    # ========================================================
    # LOCALIZAR COLUMNAS POR NOMBRE
    # ========================================================

    columnas = {}

    for celda in ws[1]:

        if celda.value is not None:

            columnas[
                str(celda.value).strip()
            ] = celda.column

    # Admitimos ambos nombres por si cambia el Excel
    if "ORDEN DE PRODUCCION (OP)" in columnas:

        col_op = columnas["ORDEN DE PRODUCCION (OP)"]

    elif "ORDEN DE PRODUCCION" in columnas:

        col_op = columnas["ORDEN DE PRODUCCION"]

    else:

        raise ValueError(
            "No se encuentra la columna "
            "'ORDEN DE PRODUCCION'."
        )

    if "UNIDADES" not in columnas:

        raise ValueError(
            "No se encuentra la columna 'UNIDADES'."
        )

    if f"{device}_PNT" not in columnas:

        if device  not in CONSOLAS:
            raise ValueError(
                "No se encuentra la columna "
                f"{device}_PNT"
            )

    col_unidades = columnas["UNIDADES"]
    # ------------------------------------------------------------
    # NUMERO DE SERIE
    # ------------------------------------------------------------

    if "NUMERO_SERIE" in columnas:

        col_ns = columnas["NUMERO_SERIE"]

    elif "NUMERO DE SERIE" in columnas:

        col_ns = columnas["NUMERO DE SERIE"]

    else:

        raise ValueError(
            "No se encuentra la columna "
            "'NUMERO_SERIE' ni 'NUMERO DE SERIE'."
        )

    if device not in CONSOLAS:
        col_pnt = columnas[f"{device}_PNT"]
    else:
        col_pnt = []


    primera_fila = 2
    ultima_fila = ws.max_row

    # ============================================================
    # FUNCIONES AUXILIARES
    # ============================================================
    
    def eliminar_columnas_vacias(ws):
        """
        Elimina las columnas completamente vacías.

        Se considera vacía una columna si, desde la fila 2 hasta
        la última fila con datos, todas sus celdas están vacías.

        La fila 1 se considera cabecera y no se utiliza para decidir
        si la columna contiene datos.
        """

        columnas_eliminar = []

        # Recorremos de derecha a izquierda para evitar
        # problemas al eliminar columnas
        for columna in range(
            ws.max_column,
            0,
            -1
        ):

            tiene_datos = False

            for fila in range(
                2,
                ws.max_row + 1
            ):

                valor = ws.cell(
                    row=fila,
                    column=columna
                ).value

                if (
                    valor is not None
                    and str(valor).strip() != ""
                ):

                    tiene_datos = True
                    break

            if not tiene_datos:

                columnas_eliminar.append(
                    columna
                )

        # Eliminar de derecha a izquierda
        for columna in columnas_eliminar:

            ws.delete_cols(
                columna,
                1
            )

    def consolidar_lotes_y_unidades_final(ws,bloques_op,pares_lote_unid):
        """
        PASADA FINAL sobre el Excel ya procesado.

        Para cada ORDEN DE PRODUCCIÓN y cada pareja:

            *_LOTE
            *_UNID

        realiza:

        1. Detecta los rangos actuales de LOTE.
        2. Si existen rangos consecutivos con el mismo lote,
        los considera un único bloque.
        3. Suma TODOS los valores de *_UNID comprendidos
        dentro del bloque del lote.
        4. Combina la celda del LOTE para todo el bloque.
        5. Combina la celda de UNIDADES para exactamente
        el mismo bloque.
        6. Nunca combina entre OP diferentes.

        Está pensada para ejecutarse al FINAL, después de las
        combinaciones anteriores.
        """

        # ========================================================
        # FUNCIONES AUXILIARES
        # ========================================================

        def es_vacio_local(valor):
            return (
                valor is None
                or str(valor).strip() == ""
            )


        def normalizar_local(valor):

            if valor is None:
                return ""

            return str(valor).strip()


        # ========================================================
        # BUSCAR MERGE VERTICAL DE UNA CELDA
        # ========================================================

        def obtener_merge_vertical(
            fila,
            columna
        ):

            for rango in ws.merged_cells.ranges:

                if (
                    rango.min_col == columna
                    and
                    rango.max_col == columna
                    and
                    rango.min_row <= fila <= rango.max_row
                ):

                    return rango

            return None


        # ========================================================
        # OBTENER SEGMENTOS ACTUALES DE LOTE
        # ========================================================

        def obtener_segmentos_lote(
            col_lote,
            inicio_op,
            final_op
        ):

            segmentos = []

            fila = inicio_op

            while fila <= final_op:

                # -----------------------------------------------
                # ¿Estamos dentro de un rango combinado?
                # -----------------------------------------------

                rango = obtener_merge_vertical(
                    fila,
                    col_lote
                )


                if rango is not None:

                    # Si estamos en una fila interior del merge,
                    # saltamos directamente al final
                    if fila != rango.min_row:

                        fila = rango.max_row + 1
                        continue


                    inicio = max(
                        rango.min_row,
                        inicio_op
                    )

                    final = min(
                        rango.max_row,
                        final_op
                    )

                    valor = ws.cell(
                        row=inicio,
                        column=col_lote
                    ).value


                    if not es_vacio_local(valor):

                        segmentos.append(
                            {
                                "inicio": inicio,
                                "final": final,
                                "valor": valor
                            }
                        )


                    fila = final + 1
                    continue


                # -----------------------------------------------
                # Celda normal
                # -----------------------------------------------

                valor = ws.cell(
                    row=fila,
                    column=col_lote
                ).value


                if not es_vacio_local(valor):

                    segmentos.append(
                        {
                            "inicio": fila,
                            "final": fila,
                            "valor": valor
                        }
                    )


                fila += 1


            return segmentos


        # ========================================================
        # SUMAR LAS UNIDADES DE UN RANGO
        # ========================================================

        def sumar_unidades_rango(
            col_unid,
            inicio,
            final
        ):

            total = 0


            for fila in range(
                inicio,
                final + 1
            ):

                celda = ws.cell(
                    row=fila,
                    column=col_unid
                )

                # Las celdas interiores de un merge no contienen
                # información real. Solo contamos las anclas.
                if isinstance(
                    celda,
                    MergedCell
                ):
                    continue


                if es_vacio_local(
                    celda.value
                ):
                    continue


                total += convertir_numero(
                    celda.value,
                    celda
                )


            return limpiar_total(
                total
            )


        # ========================================================
        # ELIMINAR MERGES DE UNA COLUMNA DENTRO DE UN RANGO
        # ========================================================

        def eliminar_merges_internos(
            columna,
            inicio,
            final
        ):

            rangos_eliminar = []


            for rango in list(
                ws.merged_cells.ranges
            ):

                if (
                    rango.min_col == columna
                    and
                    rango.max_col == columna

                    and

                    rango.min_row >= inicio
                    and
                    rango.max_row <= final
                ):

                    rangos_eliminar.append(
                        str(rango)
                    )


            for rango in rangos_eliminar:

                ws.unmerge_cells(
                    rango
                )


        # ========================================================
        # PROCESAR OP POR OP
        # ========================================================

        for (
            inicio_op,
            final_op,
            valor_op
        ) in bloques_op:


            # ====================================================
            # PROCESAR CADA PAREJA LOTE / UNID
            # ====================================================

            for (
                col_lote,
                col_unid
            ) in pares_lote_unid:


                # -----------------------------------------------
                # Obtener situación ACTUAL de la columna LOTE
                # -----------------------------------------------

                segmentos = obtener_segmentos_lote(
                    col_lote=col_lote,
                    inicio_op=inicio_op,
                    final_op=final_op
                )


                if not segmentos:
                    continue


                # =================================================
                # CONSOLIDAR SEGMENTOS CONSECUTIVOS
                # DEL MISMO LOTE
                # =================================================

                grupos = []

                grupo_actual = {
                    "inicio": segmentos[0]["inicio"],
                    "final": segmentos[0]["final"],
                    "valor": segmentos[0]["valor"]
                }


                for segmento in segmentos[1:]:

                    mismo_lote = (
                        normalizar_local(
                            segmento["valor"]
                        )
                        ==
                        normalizar_local(
                            grupo_actual["valor"]
                        )
                    )


                    consecutivo = (
                        segmento["inicio"]
                        ==
                        grupo_actual["final"] + 1
                    )


                    if (
                        mismo_lote
                        and
                        consecutivo
                    ):

                        grupo_actual["final"] = (
                            segmento["final"]
                        )


                    else:

                        grupos.append(
                            grupo_actual
                        )

                        grupo_actual = {
                            "inicio": segmento["inicio"],
                            "final": segmento["final"],
                            "valor": segmento["valor"]
                        }


                # Añadir último grupo
                grupos.append(
                    grupo_actual
                )


                # =================================================
                # PROCESAR CADA LOTE DEFINITIVO
                # =================================================

                for grupo in grupos:

                    inicio = grupo["inicio"]
                    final = grupo["final"]
                    valor_lote = grupo["valor"]


                    # ---------------------------------------------
                    # SUMAR UNIDADES ANTES DE DESHACER MERGES
                    # ---------------------------------------------

                    total_unidades = sumar_unidades_rango(
                        col_unid=col_unid,
                        inicio=inicio,
                        final=final
                    )


                    # ---------------------------------------------
                    # Eliminar merges antiguos del LOTE
                    # ---------------------------------------------

                    eliminar_merges_internos(
                        columna=col_lote,
                        inicio=inicio,
                        final=final
                    )


                    # ---------------------------------------------
                    # Eliminar merges antiguos de UNIDADES
                    # ---------------------------------------------

                    eliminar_merges_internos(
                        columna=col_unid,
                        inicio=inicio,
                        final=final
                    )


                    # ---------------------------------------------
                    # Restaurar los valores principales
                    # ---------------------------------------------

                    ws.cell(
                        row=inicio,
                        column=col_lote
                    ).value = valor_lote


                    ws.cell(
                        row=inicio,
                        column=col_unid
                    ).value = total_unidades


                    # ---------------------------------------------
                    # Crear LOS DOS merges definitivos
                    # con exactamente las mismas filas
                    # ---------------------------------------------

                    if final > inicio:

                        ws.merge_cells(
                            start_row=inicio,
                            start_column=col_lote,
                            end_row=final,
                            end_column=col_lote
                        )


                        ws.merge_cells(
                            start_row=inicio,
                            start_column=col_unid,
                            end_row=final,
                            end_column=col_unid
                        )

    def combinar_rangos_repetidos_final(ws,bloques_op):
        """
        Une rangos consecutivos que tengan el mismo valor,
        únicamente en columnas:

            *_LOTE
            *_ALBARAN
            *_PNT

        La función trabaja sobre el Excel YA FORMATEADO, por lo que
        tiene en cuenta celdas que ya están combinadas.

        Ejemplo:

            L348:L357 = "ALBARAN 1"
            L358:L362 = "ALBARAN 1"

        se convierte en:

            L348:L362 = "ALBARAN 1"

        Nunca combina entre dos órdenes de producción diferentes.
        """

        # --------------------------------------------------------
        # Normalizar valor
        # --------------------------------------------------------

        def normalizar(valor):

            if valor is None:
                return ""

            return str(valor).strip()


        # ========================================================
        # LOCALIZAR COLUMNAS A PROCESAR
        # ========================================================

        columnas_objetivo = []

        for columna in range(
            1,
            ws.max_column + 1
        ):

            cabecera = ws.cell(
                row=1,
                column=columna
            ).value

            if cabecera is None:
                continue

            cabecera_normalizada = (
                str(cabecera)
                .strip()
                .upper()
                .replace(" ", "")
            )

            if (
                cabecera_normalizada.endswith("_LOTE")
                or
                cabecera_normalizada.endswith("_ALBARAN")
                or
                cabecera_normalizada.endswith("_PNT")
            ):

                columnas_objetivo.append(
                    columna
                )


        numero_combinaciones = 0


        # ========================================================
        # PROCESAR OP POR OP
        # ========================================================

        for (inicio_op,final_op,valor_op) in bloques_op:


            # ====================================================
            # PROCESAR CADA COLUMNA
            # ====================================================

            for columna in columnas_objetivo:

                # ------------------------------------------------
                # Obtener los rangos combinados existentes
                # de ESTA columna y ESTA OP
                # ------------------------------------------------

                rangos_columna = [

                    rango

                    for rango in ws.merged_cells.ranges

                    if (
                        rango.min_col == columna
                        and
                        rango.max_col == columna
                        and
                        rango.max_row >= inicio_op
                        and
                        rango.min_row <= final_op
                    )
                ]


                rangos_columna.sort(
                    key=lambda r: r.min_row
                )


                # ------------------------------------------------
                # Crear mapa:
                #
                # fila -> rango combinado
                # ------------------------------------------------

                rango_por_fila = {}

                for rango in rangos_columna:

                    for fila in range(
                        max(rango.min_row, inicio_op),
                        min(rango.max_row, final_op) + 1
                    ):

                        rango_por_fila[
                            fila
                        ] = rango


                # =================================================
                # CONSTRUIR SEGMENTOS
                # =================================================
                #
                # Un segmento puede ser:
                #
                #   fila 350
                #
                # o
                #
                #   rango 350:360
                #
                # =================================================

                segmentos = []

                fila = inicio_op


                while fila <= final_op:

                    rango = rango_por_fila.get(
                        fila
                    )

                    # ---------------------------------------------
                    # CELDA YA COMBINADA
                    # ---------------------------------------------

                    if rango is not None:

                        # Solo procesamos desde el comienzo
                        # del rango
                        if fila != rango.min_row:

                            fila = (
                                rango.max_row + 1
                            )

                            continue


                        valor = ws.cell(
                            row=rango.min_row,
                            column=columna
                        ).value


                        segmentos.append(
                            {
                                "inicio": rango.min_row,
                                "final": rango.max_row,
                                "valor": valor
                            }
                        )


                        fila = (
                            rango.max_row + 1
                        )


                    # ---------------------------------------------
                    # CELDA NORMAL
                    # ---------------------------------------------

                    else:

                        valor = ws.cell(
                            row=fila,
                            column=columna
                        ).value


                        segmentos.append(
                            {
                                "inicio": fila,
                                "final": fila,
                                "valor": valor
                            }
                        )


                        fila += 1


                # =================================================
                # BUSCAR SEGMENTOS CONSECUTIVOS IGUALES
                # =================================================

                i = 0


                while i < len(segmentos):

                    segmento = segmentos[i]

                    valor_actual = normalizar(
                        segmento["valor"]
                    )


                    # ---------------------------------------------
                    # No agrupar vacíos
                    # ---------------------------------------------

                    if valor_actual == "":

                        i += 1
                        continue


                    inicio_grupo = (
                        segmento["inicio"]
                    )

                    final_grupo = (
                        segmento["final"]
                    )

                    valor_original = (
                        segmento["valor"]
                    )


                    j = i + 1


                    # ---------------------------------------------
                    # Buscar todos los siguientes rangos
                    # inmediatamente consecutivos y con
                    # exactamente el mismo valor
                    # ---------------------------------------------

                    while j < len(segmentos):

                        siguiente = segmentos[j]


                        # Tiene que empezar justo debajo
                        if (
                            siguiente["inicio"]
                            != final_grupo + 1
                        ):
                            break


                        # Tiene que tener el mismo valor
                        if (
                            normalizar(
                                siguiente["valor"]
                            )
                            != valor_actual
                        ):
                            break


                        # Ampliamos el grupo
                        final_grupo = (
                            siguiente["final"]
                        )

                        j += 1


                    # =================================================
                    # SI HEMOS ENCONTRADO MÁS DE UN SEGMENTO
                    # =================================================

                    if j > i + 1:


                        # ---------------------------------------------
                        # Buscar merges antiguos incluidos dentro
                        # del nuevo rango
                        # ---------------------------------------------

                        rangos_a_descombinar = []


                        for rango in list(
                            ws.merged_cells.ranges
                        ):

                            if (
                                rango.min_col == columna
                                and
                                rango.max_col == columna
                                and
                                rango.min_row >= inicio_grupo
                                and
                                rango.max_row <= final_grupo
                            ):

                                rangos_a_descombinar.append(
                                    str(rango)
                                )


                        # ---------------------------------------------
                        # Eliminar los merges pequeños
                        # ---------------------------------------------

                        for rango in rangos_a_descombinar:

                            ws.unmerge_cells(
                                rango
                            )


                        # ---------------------------------------------
                        # Restaurar valor en primera celda
                        # ---------------------------------------------

                        ws.cell(
                            row=inicio_grupo,
                            column=columna
                        ).value = valor_original


                        # ---------------------------------------------
                        # Crear merge definitivo
                        # ---------------------------------------------

                        ws.merge_cells(
                            start_row=inicio_grupo,
                            start_column=columna,
                            end_row=final_grupo,
                            end_column=columna
                        )


                        numero_combinaciones += 1


                        i = j


                    else:

                        i += 1


        # print("Rangos repetidos consolidados:",numero_combinaciones)
        
    def combinar_filas_adicionales_por_numero_serie(ws,bloques_op):
        """
        Procesa las filas adicionales generadas cuando un mismo
        NUMERO_SERIE ocupa varias filas.

        Para cada bloque de NUMERO_SERIE repetido:

        1. Combina NUMERO_SERIE.
        2. Combina SIEMPRE todas las columnas cuya cabecera
        contenga "PCB".
        3. Para el resto de columnas:
        - si la fila adicional está vacía,
            se combina con la celda superior;
        - si contiene información, se conserva.
        4. Nunca se combinan filas pertenecientes a OP diferentes.
        """

        # ========================================================
        # FUNCIONES AUXILIARES
        # ========================================================

        def es_vacio(valor):
            return (
                valor is None
                or str(valor).strip() == ""
            )


        def normalizar(valor):

            if valor is None:
                return ""

            return str(valor).strip()


        # ========================================================
        # LOCALIZAR COLUMNAS
        # ========================================================

        columnas = {}

        for celda in ws[1]:

            if celda.value is None:
                continue

            nombre = str(
                celda.value
            ).strip()

            columnas[
                nombre.upper()
            ] = celda.column


        # --------------------------------------------------------
        # NUMERO_SERIE
        # --------------------------------------------------------

        if "NUMERO_SERIE" in columnas:

            col_ns = columnas[
                "NUMERO_SERIE"
            ]

        elif "NUMERO DE SERIE" in columnas:

            col_ns = columnas[
                "NUMERO DE SERIE"
            ]

        else:

            raise ValueError(
                "No se encuentra la columna NUMERO_SERIE."
            )


        # ========================================================
        # LOCALIZAR TODAS LAS COLUMNAS PCB
        # ========================================================

        columnas_pcb = []

        for columna in range(
            1,
            ws.max_column + 1
        ):

            cabecera = ws.cell(
                row=1,
                column=columna
            ).value

            if cabecera is None:
                continue

            cabecera_normalizada = (
                str(cabecera)
                .strip()
                .upper()
                .replace(" ", "")
            )

            if "PCB" in cabecera_normalizada:

                columnas_pcb.append(
                    columna
                )


        # ========================================================
        # FUNCIÓN PARA BUSCAR MERGE EXISTENTE
        # ========================================================

        def obtener_rango_combinado(
            fila,
            columna
        ):

            for rango in list(
                ws.merged_cells.ranges
            ):

                if (
                    rango.min_row <= fila <= rango.max_row
                    and
                    rango.min_col <= columna <= rango.max_col
                ):

                    return rango

            return None


        # ========================================================
        # AMPLIAR CELDA VACÍA CON LA SUPERIOR
        # ========================================================

        def combinar_vacio_con_superior(
            fila,
            columna
        ):

            # Si la celda ya está combinada,
            # no hacemos nada
            if obtener_rango_combinado(
                fila,
                columna
            ) is not None:

                return


            fila_superior = fila - 1


            # ----------------------------------------------------
            # ¿La superior ya pertenece a un merge?
            # ----------------------------------------------------

            rango_superior = obtener_rango_combinado(
                fila_superior,
                columna
            )


            if rango_superior is not None:

                # Solo podemos ampliar si es un merge
                # vertical de ESTA columna
                if (
                    rango_superior.min_col == columna
                    and
                    rango_superior.max_col == columna
                    and
                    rango_superior.max_row == fila_superior
                ):

                    fila_inicio = (
                        rango_superior.min_row
                    )

                    valor = ws.cell(
                        row=fila_inicio,
                        column=columna
                    ).value

                    # Si arriba existe realmente información
                    if not es_vacio(valor):

                        ws.unmerge_cells(
                            str(rango_superior)
                        )

                        ws.merge_cells(
                            start_row=fila_inicio,
                            start_column=columna,
                            end_row=fila,
                            end_column=columna
                        )

                return


            # ----------------------------------------------------
            # Celda superior normal
            # ----------------------------------------------------

            celda_superior = ws.cell(
                row=fila_superior,
                column=columna
            )


            if es_vacio(
                celda_superior.value
            ):
                return


            ws.merge_cells(
                start_row=fila_superior,
                start_column=columna,
                end_row=fila,
                end_column=columna
            )


        # ========================================================
        # PRIMERO DETECTAR TODOS LOS BLOQUES DE NS REPETIDOS
        #
        # IMPORTANTE:
        # Los detectamos ANTES de empezar a hacer merges.
        # ========================================================

        bloques_ns_repetidos = []


        for (
            inicio_op,
            final_op,
            valor_op
        ) in bloques_op:


            fila = inicio_op


            while fila <= final_op:

                valor_ns = ws.cell(
                    row=fila,
                    column=col_ns
                ).value


                if es_vacio(valor_ns):

                    fila += 1
                    continue


                ns = normalizar(
                    valor_ns
                )


                inicio_ns = fila
                final_ns = fila


                siguiente = fila + 1


                while siguiente <= final_op:

                    valor_siguiente = ws.cell(
                        row=siguiente,
                        column=col_ns
                    ).value


                    if es_vacio(
                        valor_siguiente
                    ):

                        break


                    if normalizar(
                        valor_siguiente
                    ) != ns:

                        break


                    final_ns = siguiente

                    siguiente += 1


                # ------------------------------------------------
                # Solo guardar si está repetido
                # ------------------------------------------------

                if final_ns > inicio_ns:

                    bloques_ns_repetidos.append(
                        (
                            inicio_ns,
                            final_ns,
                            ns
                        )
                    )


                fila = final_ns + 1


        # ========================================================
        # PROCESAR LOS BLOQUES DETECTADOS
        # ========================================================

        for (
            inicio_ns,
            final_ns,
            ns
        ) in bloques_ns_repetidos:


            # ====================================================
            # 1. COMBINAR NUMERO_SERIE
            # ====================================================

            ws.merge_cells(
                start_row=inicio_ns,
                start_column=col_ns,
                end_row=final_ns,
                end_column=col_ns
            )


            # ====================================================
            # 2. COMBINAR SIEMPRE TODAS LAS COLUMNAS PCB
            # ====================================================

            for col_pcb in columnas_pcb:

                # -----------------------------------------------
                # Obtener el valor PCB.
                #
                # Normalmente estará en todas las filas,
                # pero buscamos el primer valor no vacío.
                # -----------------------------------------------

                valor_pcb = None


                for fila_pcb in range(
                    inicio_ns,
                    final_ns + 1
                ):

                    valor = ws.cell(
                        row=fila_pcb,
                        column=col_pcb
                    ).value


                    if not es_vacio(valor):

                        valor_pcb = valor

                        break


                # -----------------------------------------------
                # Si hubiera algún merge previo dentro de este
                # bloque PCB, lo quitamos antes de crear el
                # merge definitivo.
                # -----------------------------------------------

                rangos_a_eliminar = []


                for rango in list(
                    ws.merged_cells.ranges
                ):

                    if (
                        rango.min_col == col_pcb
                        and
                        rango.max_col == col_pcb
                        and
                        rango.min_row >= inicio_ns
                        and
                        rango.max_row <= final_ns
                    ):

                        rangos_a_eliminar.append(
                            str(rango)
                        )


                for rango in rangos_a_eliminar:

                    ws.unmerge_cells(
                        rango
                    )


                # -----------------------------------------------
                # Restaurar valor en la primera celda
                # -----------------------------------------------

                if valor_pcb is not None:

                    ws.cell(
                        row=inicio_ns,
                        column=col_pcb
                    ).value = valor_pcb


                # -----------------------------------------------
                # COMBINAR TODO EL BLOQUE DEL NS
                # -----------------------------------------------

                ws.merge_cells(
                    start_row=inicio_ns,
                    start_column=col_pcb,
                    end_row=final_ns,
                    end_column=col_pcb
                )


            # ====================================================
            # 3. RESTO DE COLUMNAS
            # ====================================================

            for fila_extra in range(
                inicio_ns + 1,
                final_ns + 1
            ):


                for columna in range(
                    1,
                    ws.max_column + 1
                ):


                    # -------------------------------------------
                    # NUMERO_SERIE ya está tratado
                    # -------------------------------------------

                    if columna == col_ns:
                        continue


                    # -------------------------------------------
                    # PCB ya está tratado
                    # -------------------------------------------

                    if columna in columnas_pcb:
                        continue


                    # -------------------------------------------
                    # Si ya está combinado,
                    # no tocarlo
                    # -------------------------------------------

                    if obtener_rango_combinado(
                        fila_extra,
                        columna
                    ) is not None:

                        continue


                    celda = ws.cell(
                        row=fila_extra,
                        column=columna
                    )


                    # -------------------------------------------
                    # Si contiene información adicional real,
                    # la dejamos tal cual
                    # -------------------------------------------

                    if not es_vacio(
                        celda.value
                    ):

                        continue


                    # -------------------------------------------
                    # Si está vacía:
                    # pertenece visualmente a la superior
                    # -------------------------------------------

                    combinar_vacio_con_superior(
                        fila=fila_extra,
                        columna=columna
                    )


        # ========================================================
        # INFORMACIÓN DE CONTROL
        # ========================================================

        # print(
        #     "Columnas PCB detectadas:",
        #     [
        #         ws.cell(
        #             row=1,
        #             column=c
        #         ).value
        #         for c in columnas_pcb
        #     ]
        # )

        # print("Bloques de NUMERO_SERIE repetido:",len(bloques_ns_repetidos))

    def es_vacio(valor):
        return (
            valor is None
            or str(valor).strip() == ""
        )

    def normalizar_valor(valor):
        """
        Normaliza un valor para comparar celdas.

        123      -> "123"
        " 123 "  -> "123"
        None     -> ""
        """

        if valor is None:
            return ""

        return str(valor).strip()

    def convertir_numero(valor, celda=None):
        """
        Convierte UNIDADES a número.

        Admite:
            1
            1.0
            "1"
            "1,5"

        Las celdas vacías se consideran 0.
        """

        if es_vacio(valor):
            return 0

        if isinstance(valor, (int, float)):
            return valor

        try:
            return float(
                str(valor)
                .strip()
                .replace(",", ".")
            )

        except ValueError:

            referencia = (
                f" en {celda.coordinate}"
                if celda is not None
                else ""
            )

            raise ValueError(
                f"El valor '{valor}'{referencia} "
                f"no puede interpretarse como unidades."
            )

    def limpiar_total(total):
        """
        Si el resultado es entero:
            5.0 -> 5
        """

        if isinstance(total, float) and total.is_integer():
            return int(total)

        return total

    def obtener_bloques_iguales(columna,fila_inicio,fila_final):
        """
        Obtiene bloques CONSECUTIVOS de valores iguales dentro
        exclusivamente del rango de filas indicado.

        Importante:
        Nunca sale fuera de la OP actual.

        Devuelve:
            [
                (fila_inicio, fila_final, valor),
                ...
            ]
        """

        bloques = []

        fila = fila_inicio

        while fila <= fila_final:

            valor = ws.cell(
                row=fila,
                column=columna
            ).value

            # No combinar celdas vacías
            if es_vacio(valor):
                fila += 1
                continue

            valor_comparacion = normalizar_valor(valor)

            inicio_bloque = fila
            final_bloque = fila

            fila += 1

            while fila <= fila_final:

                siguiente = ws.cell(
                    row=fila,
                    column=columna
                ).value

                if (
                    es_vacio(siguiente)
                    or normalizar_valor(siguiente)
                    != valor_comparacion
                ):
                    break

                final_bloque = fila
                fila += 1

            bloques.append(
                (
                    inicio_bloque,
                    final_bloque,
                    valor
                )
            )

        return bloques

    def consolidar_albaranes_pnt_con_vacios(ws,bloques_op):
        """
        Consolida las columnas *_ALBARAN y *_PNT dentro de cada OP.

        Los vacíos se consideran continuación del último valor
        no vacío.

        El bloque solamente termina cuando aparece un valor
        no vacío diferente.

        Nunca combina entre dos OP distintas.
        """

        def es_vacio(valor):
            return (
                valor is None
                or str(valor).strip() == ""
            )

        def normalizar(valor):

            if valor is None:
                return ""

            return str(valor).strip()

        # ========================================================
        # COLUMNAS ALBARAN / PNT
        # ========================================================

        columnas_objetivo = []

        for columna in range(
            1,
            ws.max_column + 1
        ):

            cabecera = ws.cell(
                row=1,
                column=columna
            ).value

            if cabecera is None:
                continue

            nombre = (
                str(cabecera)
                .strip()
                .upper()
                .replace(" ", "")
            )

            if (
                nombre.endswith("_ALBARAN")
                or
                nombre.endswith("_PNT")
            ):

                columnas_objetivo.append(
                    columna
                )

        # ========================================================
        # PROCESAR OP POR OP
        # ========================================================

        for inicio_op, final_op, valor_op in bloques_op:

            for columna in columnas_objetivo:

                # ------------------------------------------------
                # Primero descombinar ESTA columna dentro de la OP
                # conservando los valores
                # ------------------------------------------------

                rangos = []

                for rango in list(
                    ws.merged_cells.ranges
                ):

                    if (
                        rango.min_col == columna
                        and
                        rango.max_col == columna
                        and
                        rango.min_row >= inicio_op
                        and
                        rango.max_row <= final_op
                    ):

                        valor = ws.cell(
                            row=rango.min_row,
                            column=columna
                        ).value

                        rangos.append(
                            (
                                str(rango),
                                rango.min_row,
                                valor
                            )
                        )

                # -----------------------------------------------
                # Descombinar y restaurar valor superior
                # -----------------------------------------------

                for (
                    nombre_rango,
                    fila_inicio,
                    valor
                ) in rangos:

                    ws.unmerge_cells(
                        nombre_rango
                    )

                # =================================================
                # RECORRER LA COLUMNA
                # =================================================

                fila = inicio_op

                while fila <= final_op:

                    valor = ws.cell(
                        row=fila,
                        column=columna
                    ).value

                    # No podemos empezar un bloque con vacío
                    if es_vacio(valor):

                        fila += 1
                        continue

                    valor_actual = normalizar(
                        valor
                    )

                    inicio_bloque = fila
                    final_bloque = fila

                    siguiente = fila + 1

                    # =============================================
                    # AMPLIAR HASTA ENCONTRAR VALOR DIFERENTE
                    # =============================================

                    while siguiente <= final_op:

                        valor_siguiente = ws.cell(
                            row=siguiente,
                            column=columna
                        ).value

                        # VACÍO:
                        # continúa aplicando el valor anterior
                        if es_vacio(
                            valor_siguiente
                        ):

                            final_bloque = siguiente
                            siguiente += 1
                            continue

                        # MISMO VALOR:
                        # continúa el mismo bloque
                        if (
                            normalizar(valor_siguiente)
                            == valor_actual
                        ):

                            final_bloque = siguiente
                            siguiente += 1
                            continue

                        # Valor real diferente
                        break

                    # =============================================
                    # CREAR MERGE DEFINITIVO
                    # =============================================

                    if final_bloque > inicio_bloque:

                        ws.merge_cells(
                            start_row=inicio_bloque,
                            start_column=columna,
                            end_row=final_bloque,
                            end_column=columna
                        )

                    fila = siguiente

    def aplicar_bordes_contenido(ws):

        borde_fino = Side(style="thin", color="000000")

        def es_vacio(valor):
            if valor is None:
                return True
            if isinstance(valor, str):
                return valor.strip() == ""
            return False

        def aplicar_lados(celda, izquierda=None, derecha=None, arriba=None, abajo=None):
            celda.border = Border(left=izquierda if izquierda is not None else celda.border.left, right=derecha if derecha is not None else celda.border.right, top=arriba if arriba is not None else celda.border.top, bottom=abajo if abajo is not None else celda.border.bottom)

        # ========================================================
        # GUARDAR TODAS LAS CELDAS QUE PERTENECEN A MERGES
        # ========================================================

        celdas_combinadas = set()

        for rango in ws.merged_cells.ranges:
            for fila in range(rango.min_row, rango.max_row + 1):
                for columna in range(rango.min_col, rango.max_col + 1):
                    celdas_combinadas.add((fila, columna))

        # ========================================================
        # CELDAS NORMALES NO VACÍAS
        # ========================================================

        for fila in ws.iter_rows():
            for celda in fila:

                if (celda.row, celda.column) in celdas_combinadas:
                    continue

                if es_vacio(celda.value):
                    continue

                celda.border = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

        # ========================================================
        # RANGOS COMBINADOS
        # Solo dibujamos el CONTORNO EXTERIOR
        # ========================================================

        for rango in ws.merged_cells.ranges:

            celda_principal = ws.cell(rango.min_row, rango.min_col)

            if es_vacio(celda_principal.value):
                continue

            # Borde superior
            for columna in range(rango.min_col, rango.max_col + 1):
                aplicar_lados(ws.cell(rango.min_row, columna), arriba=borde_fino)

            # Borde inferior
            for columna in range(rango.min_col, rango.max_col + 1):
                aplicar_lados(ws.cell(rango.max_row, columna), abajo=borde_fino)

            # Borde izquierdo
            for fila in range(rango.min_row, rango.max_row + 1):
                aplicar_lados(ws.cell(fila, rango.min_col), izquierda=borde_fino)

            # Borde derecho
            for fila in range(rango.min_row, rango.max_row + 1):
                aplicar_lados(ws.cell(fila, rango.max_col), derecha=borde_fino)

    # ========================================================
    # ELIMINAMOS LAS COLUMNAS VACIAS
    # ========================================================
    # MUY IMPORTANTE HACER AL PRINCIPIO YA QUE AFECTA A LA COMBINACION DE CELDAS
    # eliminar_columnas_vacias(ws=ws)

    # ============================================================
    # IDENTIFICAR COLUMNAS
    # ============================================================

    columnas = {}

    for celda in ws[1]:

        if celda.value is None:
            continue

        nombre = str(
            celda.value
        ).strip()

        columnas[nombre] = celda.column


    # ------------------------------------------------------------
    # ORDEN DE PRODUCCIÓN
    # ------------------------------------------------------------

    if "ORDEN DE PRODUCCION (OP)" in columnas:

        col_op = columnas[
            "ORDEN DE PRODUCCION (OP)"
        ]

    elif "ORDEN DE PRODUCCION" in columnas:

        col_op = columnas[
            "ORDEN DE PRODUCCION"
        ]

    else:

        raise ValueError(
            "No se encuentra la columna "
            "'ORDEN DE PRODUCCION'."
        )


    # ------------------------------------------------------------
    # UNIDADES GENERALES DE LA OP
    # ------------------------------------------------------------

    if "UNIDADES" not in columnas:

        raise ValueError(
            "No se encuentra la columna 'UNIDADES'."
        )

    col_unidades = columnas["UNIDADES"]

    # ============================================================
    # IDENTIFICAR AUTOMÁTICAMENTE TODAS LAS COLUMNAS
    # *_LOTE, *_UNID Y *_PNT
    # ============================================================

    columnas_lote = []
    columnas_unid = []
    columnas_albaran = []
    columnas_pnt = []

    for numero_columna in range(1, ws.max_column + 1):

        cabecera = ws.cell(row=1,column=numero_columna).value

        if cabecera is None:
            continue

        cabecera_normalizada = (
            str(cabecera)
            .strip()
            .upper()
            .replace(" ", "")
        )

        if cabecera_normalizada.endswith("_LOTE"):

            columnas_lote.append(
                numero_columna
            )

        elif cabecera_normalizada.endswith("_UNID"):

            columnas_unid.append(
                numero_columna
            )

        elif cabecera_normalizada.endswith("_ALBARAN"):

            columnas_albaran.append(
                numero_columna
            )

        elif cabecera_normalizada.endswith("_PNT"):

            columnas_pnt.append(
                numero_columna
            )


    # ============================================================
    # COMPROBAR RELACIÓN LOTE -> UNID
    # ============================================================

    pares_lote_unid = []

    for col_lote in columnas_lote:

        col_posible_unid = col_lote + 1

        if col_posible_unid > ws.max_column:
            continue

        cabecera_unid = ws.cell(
            row=1,
            column=col_posible_unid
        ).value

        if cabecera_unid is None:
            continue

        cabecera_unid_normalizada = (
            str(cabecera_unid)
            .strip()
            .upper()
            .replace(" ", "")
        )

        if cabecera_unid_normalizada.endswith(
            "_UNID"
        ):

            pares_lote_unid.append(
                (
                    col_lote,
                    col_posible_unid
                )
            )

    # ============================================================
    # OBTENER BLOQUES DE ÓRDENES DE PRODUCCIÓN
    # ============================================================

    primera_fila = 2
    ultima_fila = ws.max_row

    bloques_op = []

    fila = primera_fila

    while fila <= ultima_fila:

        valor_op = ws.cell(
            row=fila,
            column=col_op
        ).value

        # Ignorar filas sin OP
        if es_vacio(valor_op):

            fila += 1
            continue

        op_comparacion = normalizar_valor(
            valor_op
        )

        inicio_op = fila
        final_op = fila

        fila += 1


        while fila <= ultima_fila:

            siguiente_op = ws.cell(
                row=fila,
                column=col_op
            ).value

            if (
                es_vacio(siguiente_op)
                or normalizar_valor(siguiente_op)
                != op_comparacion
            ):
                break

            final_op = fila

            fila += 1


        bloques_op.append(
            (
                inicio_op,
                final_op,
                valor_op
            )
        )


    # ============================================================
    # PROCESAR CADA ORDEN DE PRODUCCIÓN POR SEPARADO
    # ============================================================

    for inicio_op, final_op, valor_op in bloques_op:

        # ========================================================
        # 1. SUMAR UNIDADES GENERALES DE LA OP
        # ========================================================

        total_unidades_op = 0


        # ========================================================
        # 1. CALCULAR UNIDADES DE LA OP
        #    = NÚMEROS DE SERIE ÚNICOS
        # ========================================================

        numeros_serie_op = set()

        for fila in range(
            inicio_op,
            final_op + 1
        ):

            numero_serie = ws.cell(
                row=fila,
                column=col_ns
            ).value

            # Ignorar celdas vacías
            if es_vacio(numero_serie):
                continue

            # Normalizar para evitar diferencias por espacios
            numero_serie = normalizar_valor(
                numero_serie
            )

            numeros_serie_op.add(
                numero_serie
            )


        total_unidades_op = len(
            numeros_serie_op
        )


        # Escribir resultado ANTES de combinar
        ws.cell(
            row=inicio_op,
            column=col_unidades
        ).value = total_unidades_op


        # ========================================================
        # 2. COMBINAR ORDEN DE PRODUCCIÓN Y UNIDADES
        # ========================================================

        if final_op > inicio_op:

            ws.merge_cells(
                start_row=inicio_op,
                start_column=col_op,
                end_row=final_op,
                end_column=col_op
            )

            ws.merge_cells(
                start_row=inicio_op,
                start_column=col_unidades,
                end_row=final_op,
                end_column=col_unidades
            )


        # Centrar
        ws.cell(
            row=inicio_op,
            column=col_op
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False
        )

        ws.cell(
            row=inicio_op,
            column=col_unidades
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False
        )


        # ========================================================
        # 3. COLUMNAS *_LOTE + *_UNID
        # ========================================================

        for col_lote, col_unid in pares_lote_unid:

            # ----------------------------------------------------
            # Localizar bloques del MISMO LOTE
            # exclusivamente dentro de esta OP
            # ----------------------------------------------------

            bloques_lote = obtener_bloques_iguales(
                columna=col_lote,
                fila_inicio=inicio_op,
                fila_final=final_op
            )


            for (
                inicio_lote,
                final_lote,
                valor_lote
            ) in bloques_lote:

                # =================================================
                # SUMAR UNIDADES DEL LOTE
                # =================================================

                total_unidades_lote = 0


                for fila_lote in range(
                    inicio_lote,
                    final_lote + 1
                ):

                    celda_unidad = ws.cell(
                        row=fila_lote,
                        column=col_unid
                    )

                    total_unidades_lote += (
                        convertir_numero(
                            celda_unidad.value,
                            celda_unidad
                        )
                    )


                total_unidades_lote = limpiar_total(
                    total_unidades_lote
                )


                # Escribir total
                ws.cell(
                    row=inicio_lote,
                    column=col_unid
                ).value = total_unidades_lote


                # =================================================
                # COMBINAR LOTE Y UNIDADES
                # =================================================

                if final_lote > inicio_lote:

                    # Combinar LOTE
                    ws.merge_cells(
                        start_row=inicio_lote,
                        start_column=col_lote,
                        end_row=final_lote,
                        end_column=col_lote
                    )

                    # Combinar UNIDADES
                    ws.merge_cells(
                        start_row=inicio_lote,
                        start_column=col_unid,
                        end_row=final_lote,
                        end_column=col_unid
                    )


                # =================================================
                # ALINEACIÓN
                # =================================================

                ws.cell(
                    row=inicio_lote,
                    column=col_lote
                ).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False
                )

                ws.cell(
                    row=inicio_lote,
                    column=col_unid
                ).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False
                )

        # ========================================================
        # 4. COLUMNAS *_ALBARAN
        # ========================================================

        for col_albaran in columnas_albaran:

            # Buscar albaranes iguales únicamente dentro
            # de la ORDEN DE PRODUCCIÓN actual
            bloques_albaran = obtener_bloques_iguales(
                columna=col_albaran,
                fila_inicio=inicio_op,
                fila_final=final_op
            )

            for (
                inicio_albaran,
                final_albaran,
                valor_albaran
            ) in bloques_albaran:

                # Combinar solamente si ocupa más de una fila
                if final_albaran > inicio_albaran:

                    ws.merge_cells(
                        start_row=inicio_albaran,
                        start_column=col_albaran,
                        end_row=final_albaran,
                        end_column=col_albaran
                    )

                # Centrar el valor del albarán
                ws.cell(
                    row=inicio_albaran,
                    column=col_albaran
                ).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False
                )
        # ========================================================
        # 5. COLUMNAS *_PNT
        # ========================================================

        for col_pnt in columnas_pnt:

            bloques_pnt = obtener_bloques_iguales(
                columna=col_pnt,
                fila_inicio=inicio_op,
                fila_final=final_op
            )


            for (
                inicio_pnt,
                final_pnt,
                valor_pnt
            ) in bloques_pnt:

                if final_pnt > inicio_pnt:

                    ws.merge_cells(
                        start_row=inicio_pnt,
                        start_column=col_pnt,
                        end_row=final_pnt,
                        end_column=col_pnt
                    )


                ws.cell(
                    row=inicio_pnt,
                    column=col_pnt
                ).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False
                )

    # ========================================================
    # CORREGIR FILAS ADICIONALES GENERADAS POR
    # COMPONENTES CON MÁS DE UN LOTE
    # ========================================================

    combinar_filas_adicionales_por_numero_serie(ws=ws,bloques_op=bloques_op)

    # ========================================================
    # CONSOLIDAR RANGOS REPETIDOS
    # ========================================================

    combinar_rangos_repetidos_final(ws=ws,bloques_op=bloques_op)

    # ========================================================
    # CONSOLIDAR LOTES Y UNIDADES FINAL
    # ========================================================

    consolidar_lotes_y_unidades_final(ws=ws, bloques_op=bloques_op, pares_lote_unid=pares_lote_unid)

    # ========================================================
    # CONSOLIDAR ALBARANES Y PNT VACIOS
    # ========================================================  
    
    consolidar_albaranes_pnt_con_vacios(ws=ws,bloques_op=bloques_op)

    # ========================================================
    # FORMATO GENERAL DE LA HOJA
    # ========================================================

    # --------------------------------------------------------
    # 1. CABECERA EN NEGRITA Y SIN SALTO DE LÍNEA
    # --------------------------------------------------------

    for celda in ws[1]:

        if celda.value is not None:

            # Eliminar posibles saltos de línea existentes
            if isinstance(celda.value, str):
                celda.value = (
                    celda.value
                    .replace("\n", " ")
                    .replace("\r", " ")
                )

            celda.font = Font(
                name=celda.font.name or "Calibri",
                size=celda.font.sz or 11,
                bold=True
            )

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False
            )

    # --------------------------------------------------------
    # 2. EVITAR SALTOS DE TEXTO EN TODAS LAS CELDAS
    # --------------------------------------------------------

    for fila in ws.iter_rows():

        for celda in fila:

            # Las celdas internas de un rango combinado no
            # contienen datos y no necesitamos modificarlas
            if celda.value is None:
                continue

            # Eliminar saltos de línea reales
            if isinstance(celda.value, str):

                celda.value = (
                    celda.value
                    .replace("\n", " ")
                    .replace("\r", " ")
                )

            # Conservar alineación horizontal existente,
            # pero impedir que Excel divida el texto
            celda.alignment = Alignment(
                horizontal=celda.alignment.horizontal,
                vertical="center",
                wrap_text=False
            )


    # --------------------------------------------------------
    # 3. AJUSTAR AUTOMÁTICAMENTE ANCHO DE COLUMNAS
    # --------------------------------------------------------

    for numero_columna in range(1,ws.max_column + 1):

        longitud_maxima = 0

        letra_columna = get_column_letter(
            numero_columna
        )

        for numero_fila in range(
            1,
            ws.max_row + 1
        ):

            celda = ws.cell(
                row=numero_fila,
                column=numero_columna
            )

            valor = celda.value

            if valor is None:
                continue

            texto = str(valor)

            # En caso de que haya quedado algún salto
            texto = (
                texto
                .replace("\n", " ")
                .replace("\r", " ")
            )

            longitud = len(texto)

            if longitud > longitud_maxima:
                longitud_maxima = longitud

        # ----------------------------------------------------
        # Añadir margen para que el texto no quede pegado
        # ----------------------------------------------------

        ancho = longitud_maxima + 3

        ws.column_dimensions[
            letra_columna
        ].width = ancho

    # --------------------------------------------------------
    # 4. CENTRAR TODAS LAS CELDAS
    # --------------------------------------------------------

    alineacion_centrada = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False
    )

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.alignment = alineacion_centrada


    # --------------------------------------------------------
    # 4. APLICAMOS EL FORMATO FINAL PARA QUE APAREZCAN LOS BORDES
    # --------------------------------------------------------
    aplicar_bordes_contenido(ws)
    # ========================================================
    # GUARDAR
    # ========================================================

    wb.save(
        ruta_salida
    )

    return ruta_salida

def revisa_informacion_bachrecord(df, dispositivo):
    """
    Revisa la información necesaria para generar el Batch Record.

    Tiene en cuenta que un mismo NUMERO_SERIE puede ocupar
    varias filas porque algún componente utiliza más de un lote.

    Para un mismo NUMERO_SERIE:
        - una celda vacía NO se considera error si en otra fila
          del mismo NS esa columna tiene información.
        - solo se considera información faltante si TODAS las
          filas de ese NS están vacías para esa columna.

    Devuelve una lista de incidencias.
    """

    info_faltante = []

    # ========================================================
    # FUNCIÓN INTERNA
    # ========================================================

    # def revisar_celdas_vacias(df_articulo,articulo):

    #     faltantes = []

    #     columna_ns = "NUMERO_SERIE"

    #     if columna_ns not in df_articulo.columns:

    #         raise ValueError(
    #             f"No se encuentra la columna "
    #             f"'{columna_ns}'"
    #         )

    #     # ====================================================
    #     # SABER SI UN VALOR ESTÁ VACÍO
    #     # ====================================================

    #     def es_vacio(valor):

    #         if pd.isna(valor):
    #             return True

    #         if isinstance(valor, str):
    #             return valor.strip() == ""

    #         return False


    #     # ====================================================
    #     # AGRUPAR FILAS CONSECUTIVAS DEL MISMO NS
    #     # ====================================================

    #     bloques_ns = []

    #     fila = 0

    #     while fila < len(df_articulo):

    #         numero_serie = df_articulo.iloc[
    #             fila
    #         ][columna_ns]

    #         # ------------------------------------------------
    #         # Si falta directamente el número de serie,
    #         # esto sí es un error real
    #         # ------------------------------------------------

    #         if es_vacio(numero_serie):

    #             faltantes.append(
    #                 {
    #                     "msg": (
    #                         f"En el excel resumen del {articulo} "
    #                         f"se ha encontrado una fila sin "
    #                         f"NUMERO_SERIE"
    #                     ),
    #                     "columna": columna_ns,
    #                     "lote_principal": None
    #                 }
    #             )

    #             fila += 1
    #             continue


    #         numero_serie = str(
    #             numero_serie
    #         ).strip()

    #         inicio = fila
    #         final = fila

    #         siguiente = fila + 1


    #         # ------------------------------------------------
    #         # Buscar filas consecutivas del mismo NS
    #         # ------------------------------------------------

    #         while siguiente < len(df_articulo):

    #             siguiente_ns = df_articulo.iloc[
    #                 siguiente
    #             ][columna_ns]

    #             if es_vacio(siguiente_ns):
    #                 break

    #             if (
    #                 str(siguiente_ns).strip()
    #                 != numero_serie
    #             ):
    #                 break

    #             final = siguiente
    #             siguiente += 1


    #         bloques_ns.append(
    #             (
    #                 inicio,
    #                 final,
    #                 numero_serie
    #             )
    #         )

    #         fila = final + 1


    #     # ====================================================
    #     # REVISAR CADA BLOQUE DE NS
    #     # ====================================================

    #     for (
    #         inicio,
    #         final,
    #         numero_serie
    #     ) in bloques_ns:

    #         # Filas correspondientes al mismo dispositivo
    #         bloque = df_articulo.iloc[
    #             inicio:final + 1
    #         ]


    #         # =================================================
    #         # REVISAR CADA COLUMNA
    #         # =================================================

    #         for columna in df_articulo.columns:

    #             # NUMERO_SERIE ya está comprobado
    #             if columna == columna_ns:
    #                 continue


    #             # ---------------------------------------------
    #             # Obtener todos los valores de esta columna
    #             # para este mismo número de serie
    #             # ---------------------------------------------

    #             valores = bloque[
    #                 columna
    #             ].tolist()


    #             # ---------------------------------------------
    #             # SOLO hay error si TODOS están vacíos
    #             # ---------------------------------------------

    #             todos_vacios = all(
    #                 es_vacio(valor)
    #                 for valor in valores
    #             )


    #             if todos_vacios:

    #                 faltantes.append(
    #                     {
    #                         "msg": (
    #                             f"En el excel resumen del "
    #                             f"{articulo} falta información "
    #                             f"en la columna {columna} "
    #                             f"del artículo {numero_serie}"
    #                         ),
    #                         "columna": columna,
    #                         "lote_principal": numero_serie
    #                     }
    #                 )


    #     return faltantes

    def revisar_celdas_vacias(df_articulo, articulo):

        faltantes = []
        columna_ns = "NUMERO_SERIE"

        if columna_ns not in df_articulo.columns:
            raise ValueError(f"No se encuentra la columna '{columna_ns}'")

        # ========================================================
        # SABER SI UN VALOR ESTÁ VACÍO
        # ========================================================

        def es_vacio(valor):
            if pd.isna(valor):
                return True
            if isinstance(valor, str):
                return valor.strip() == ""
            return False

        # ========================================================
        # LOCALIZAR GRUPOS DE 4 COLUMNAS DE CADA ARTÍCULO
        # ========================================================

        sufijos_articulo = ("_LOTE", "_UNID", "_ALBARAN", "_PNT")
        grupos_articulos = {}

        for columna in df_articulo.columns:
            nombre = str(columna).strip()

            for sufijo in sufijos_articulo:
                if nombre.upper().endswith(sufijo):
                    nombre_articulo = nombre[:-len(sufijo)]
                    grupos_articulos.setdefault(nombre_articulo, {})[sufijo] = columna
                    break

        # ========================================================
        # COMPROBAR SI UN ARTÍCULO ESTÁ COMPLETAMENTE VACÍO
        # PARA UN DETERMINADO BLOQUE DE NUMERO_SERIE
        # ========================================================

        def articulo_completamente_vacio(bloque, nombre_articulo):

            columnas = grupos_articulos.get(nombre_articulo, {})

            # Solo aplicamos esta excepción si existen las 4 columnas
            if not all(sufijo in columnas for sufijo in sufijos_articulo):
                return False

            for sufijo in sufijos_articulo:
                columna = columnas[sufijo]

                for valor in bloque[columna]:
                    if not es_vacio(valor):
                        return False

            return True

        # ========================================================
        # AGRUPAR FILAS CONSECUTIVAS DEL MISMO NS
        # ========================================================

        bloques_ns = []
        fila = 0

        while fila < len(df_articulo):

            numero_serie = df_articulo.iloc[fila][columna_ns]

            if es_vacio(numero_serie):

                faltantes.append({
                    "msg": f"En el excel resumen del {articulo} se ha encontrado una fila sin NUMERO_SERIE",
                    "columna": columna_ns,
                    "lote_principal": None
                })

                fila += 1
                continue

            numero_serie = str(numero_serie).strip()
            inicio = fila
            final = fila
            siguiente = fila + 1

            while siguiente < len(df_articulo):

                siguiente_ns = df_articulo.iloc[siguiente][columna_ns]

                if es_vacio(siguiente_ns):
                    break

                if str(siguiente_ns).strip() != numero_serie:
                    break

                final = siguiente
                siguiente += 1

            bloques_ns.append((inicio, final, numero_serie))
            fila = final + 1

        # ========================================================
        # REVISAR CADA BLOQUE DE NS
        # ========================================================

        for inicio, final, numero_serie in bloques_ns:

            bloque = df_articulo.iloc[inicio:final + 1]

            # ----------------------------------------------------
            # ARTÍCULOS QUE NO APLICAN A ESTE NUMERO_SERIE
            #
            # Si LOTE + UNID + ALBARAN + PNT están todos vacíos,
            # se considera correcto y no se revisan esas columnas.
            # ----------------------------------------------------

            articulos_vacios = set()

            for nombre_articulo in grupos_articulos:
                if articulo_completamente_vacio(bloque, nombre_articulo):
                    articulos_vacios.add(nombre_articulo)

            # ====================================================
            # REVISAR CADA COLUMNA
            # ====================================================

            for columna in df_articulo.columns:

                if columna == columna_ns:
                    continue

                nombre_columna = str(columna).strip()
                articulo_columna = None

                # ------------------------------------------------
                # SABER SI LA COLUMNA PERTENECE A UN ARTÍCULO
                # ------------------------------------------------

                for sufijo in sufijos_articulo:
                    if nombre_columna.upper().endswith(sufijo):
                        articulo_columna = nombre_columna[:-len(sufijo)]
                        break

                # ------------------------------------------------
                # SI LAS 4 COLUMNAS DE ESE ARTÍCULO ESTÁN VACÍAS,
                # NO ES UN ERROR: EL ARTÍCULO NO APLICA
                # ------------------------------------------------

                if articulo_columna in articulos_vacios:
                    continue

                # ------------------------------------------------
                # PARA EL RESTO:
                # SOLO ERROR SI TODAS LAS FILAS DEL MISMO NS
                # ESTÁN VACÍAS EN ESA COLUMNA
                # ------------------------------------------------

                valores = bloque[columna].tolist()
                todos_vacios = all(es_vacio(valor) for valor in valores)

                if todos_vacios:

                    faltantes.append({
                        "msg": f"En el excel resumen del {articulo} falta información en la columna {columna} del artículo {numero_serie}",
                        "columna": columna,
                        "lote_principal": numero_serie
                    })

        return faltantes

    # ========================================================
    # EJECUTAR REVISIÓN
    # ========================================================

    info_faltante.extend(revisar_celdas_vacias(df,f"Articulo {dispositivo}" ))

    return info_faltante

def localizar_documentacion_batchrecord(dispositivo_br, consola_br, carpeta_documentos=RUTA_ERP, carpeta_pnt=RUTA_PNT, extensiones=(".pdf", ".doc", ".docx", ".xlsx", ".xls", ".xlsm")):

    def cargar_dataframe(fuente):
        if isinstance(fuente, pd.DataFrame):
            return fuente.copy()
        return pd.read_excel(fuente, dtype=object)

    def es_vacio(valor):
        if pd.isna(valor):
            return True
        if isinstance(valor, str):
            return valor.strip() == ""
        return False

    def normalizar_valor(valor):
        if es_vacio(valor):
            return None
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        return str(valor).strip()

    def valores_unicos(df, columna):
        if columna not in df.columns:
            return []

        resultado = []
        vistos = set()

        for valor in df[columna]:
            valor = normalizar_valor(valor)

            if valor is None or valor in vistos:
                continue

            vistos.add(valor)
            resultado.append(valor)

        return resultado

    def normalizar_texto(texto):
        if texto is None:
            return ""

        texto = unicodedata.normalize("NFKD", str(texto))
        texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
        return re.sub(r"[^A-Z0-9]", "", texto.upper())

    def quitar_fecha(referencia):
        referencia = str(referencia).strip()
        return re.sub(r"\s+\d{1,2}[\/_-]\d{1,2}[\/_-]\d{2,4}$", "", referencia).strip()

    def extension_valida(archivo):
        extensiones_validas = {ext.lower() if str(ext).startswith(".") else f".{str(ext).lower()}" for ext in extensiones}
        return archivo.suffix.lower() in extensiones_validas

    def obtener_proveedor_ruta(ruta):
        try:
            ruta_relativa = Path(ruta).resolve().relative_to(Path(carpeta_documentos).resolve())
            return ruta_relativa.parts[0] if ruta_relativa.parts else None
        except (ValueError, OSError):
            return None

    # ========================================================
    # INDEXAR DOCUMENTOS ERP
    # ========================================================

    def indexar_carpeta(carpeta):
        carpeta = Path(carpeta)

        if not carpeta.exists():
            raise FileNotFoundError(f"No existe la carpeta: {carpeta}")

        archivos = []

        for archivo in carpeta.rglob("*"):
            if not archivo.is_file() or not extension_valida(archivo):
                continue

            archivos.append({
                "ruta": str(archivo),
                "nombre": archivo.name,
                "stem": archivo.stem,
                "stem_normalizado": normalizar_texto(archivo.stem)
            })

        return archivos

    df_dispositivo = cargar_dataframe(dispositivo_br)
    df_consola = cargar_dataframe(consola_br)

    archivos_documentos = indexar_carpeta(carpeta_documentos)
    archivos_pnt = indexar_carpeta(carpeta_pnt)

    # ========================================================
    # LOCALIZAR CARPETA EXACTA DE I+D PARA UN ARTÍCULO
    # ========================================================

    def obtener_carpeta_id_articulo(articulo):
        carpeta_id = Path(carpeta_documentos) / "I+D"

        if not carpeta_id.exists():
            return None

        carpeta_directa = carpeta_id / articulo

        if carpeta_directa.exists():
            return carpeta_directa

        # Búsqueda sin distinguir mayúsculas/minúsculas
        for carpeta in carpeta_id.iterdir():
            if carpeta.is_dir() and carpeta.name.strip().upper() == str(articulo).strip().upper():
                return carpeta

        return None

    # ========================================================
    # BUSCAR ORDEN / OP / ALBARÁN
    # ========================================================

    def buscar_albaran_op(referencia, tipo, articulo=None):
        referencia = normalizar_valor(referencia)

        if referencia is None:
            return []

        # ====================================================
        # ORDEN INTERNA / OP
        #
        # Siempre están en:
        #
        # ERP\I+D\<ARTICULO>\
        #
        # Se busca únicamente el número inicial.
        # ====================================================

        if tipo in ("ORDEN", "OP"):

            if articulo is None:
                return []

            carpeta_articulo = obtener_carpeta_id_articulo(articulo)

            if carpeta_articulo is None:
                return []

            if tipo == "ORDEN":
                coincidencia = re.search(r"1//\s*(\d+)", referencia)
            else:
                coincidencia = re.match(r"\s*(\d+)", referencia)

            if not coincidencia:
                return []

            numero = coincidencia.group(1)
            patron = re.compile(rf"^\s*{re.escape(numero)}(?!\d)")
            rutas = []

            for archivo in carpeta_articulo.rglob("*"):
                if not archivo.is_file() or not extension_valida(archivo):
                    continue

                if patron.search(archivo.stem):
                    rutas.append(str(archivo))

            return list(dict.fromkeys(rutas))

        # ====================================================
        # ALBARÁN
        #
        # Buscar únicamente sobre el NOMBRE DEL FICHERO.
        # No sobre toda la ruta.
        # ====================================================

        referencia_reducida = quitar_fecha(referencia)
        referencia_normalizada = normalizar_texto(referencia_reducida)

        if not referencia_normalizada:
            return []

        rutas = []

        for archivo in archivos_documentos:
            if referencia_normalizada in archivo["stem_normalizado"]:
                rutas.append(archivo["ruta"])

        return list(dict.fromkeys(rutas))

    # ========================================================
    # BUSCAR PNT
    # ========================================================

    def buscar_pnt(referencia):
        referencia = normalizar_valor(referencia)

        if referencia is None:
            return []

        rutas = []

        if referencia.isdigit():
            patron = re.compile(rf"(?<!\d){re.escape(referencia)}(?!\d)")

            for archivo in archivos_pnt:
                if patron.search(archivo["stem"]):
                    rutas.append(archivo["ruta"])

        else:
            referencia_normalizada = normalizar_texto(referencia)

            for archivo in archivos_pnt:
                if referencia_normalizada in archivo["stem_normalizado"]:
                    rutas.append(archivo["ruta"])

        return list(dict.fromkeys(rutas))

    # ========================================================
    # BUSCAR RESUMEN LOTES
    # ========================================================

    def localizar_resumenes_articulos(estructura_documental, carpeta_erp=RUTA_ERP):
        carpeta_erp = Path(carpeta_erp)
        carpeta_id = carpeta_erp / "I+D"
        extensiones_excel = {".xlsx", ".xls", ".xlsm", ".xlsb"}

        if not carpeta_erp.exists():
            raise FileNotFoundError(f"No existe la carpeta ERP: {carpeta_erp}")

        def pertenece_a_carpeta_articulo(archivo, articulo, carpeta_base):
            articulo_normalizado = str(articulo).strip().upper()

            try:
                relativa = archivo.relative_to(carpeta_base)
            except ValueError:
                return False

            return any(parte.strip().upper() == articulo_normalizado for parte in relativa.parts[:-1])

        for articulo, datos_articulo in estructura_documental.get("componentes", {}).items():

            articulo_normalizado = str(articulo).strip().upper()

            pdf_lotes = []
            excel_lotes = []

            pdf_resumen_id = []
            excel_resumen_id = []

            # ========================================================
            # 1. BÚSQUEDA NORMAL EN ERP
            #    EXCLUYENDO I+D
            #
            #    Nombre:
            #        ARTICULO + LOTES
            # ========================================================

            for archivo in carpeta_erp.rglob("*"):

                if not archivo.is_file():
                    continue

                if archivo.name.startswith("~$"):
                    continue

                try:
                    relativa_erp = archivo.relative_to(carpeta_erp)
                except ValueError:
                    continue

                # No buscar aquí dentro de I+D
                if relativa_erp.parts and relativa_erp.parts[0].strip().upper() == "I+D":
                    continue

                if not pertenece_a_carpeta_articulo(archivo, articulo, carpeta_erp):
                    continue

                nombre_normalizado = archivo.stem.strip().upper()

                if articulo_normalizado not in nombre_normalizado:
                    continue

                if "LOTES" not in nombre_normalizado:
                    continue

                if archivo.suffix.lower() == ".pdf":
                    pdf_lotes.append(str(archivo))

                elif archivo.suffix.lower() in extensiones_excel:
                    excel_lotes.append(str(archivo))

            # ========================================================
            # SI HEMOS ENCONTRADO PDF DE LOTES -> TERMINAMOS
            # ========================================================

            if pdf_lotes:
                datos_articulo["resumen"] = pdf_lotes[0]

                # print(f"RESUMEN LOTES -> {articulo}: {pdf_lotes[0]}")

                continue

            # ========================================================
            # 2. FALLBACK EN ERP\I+D
            #
            #    Buscar:
            #        ARTICULO_RESUMEN...
            #
            #    Ejemplos admitidos:
            #        CABLEPLANO_RESUMEN.pdf
            #        CABLEPLANO_RESUMEN_formateado.pdf
            #        CABLEPLANO_RESUMEN.xlsx
            # ========================================================

            if carpeta_id.exists():

                prefijo_resumen = f"{articulo_normalizado}_RESUMEN"

                for archivo in carpeta_id.rglob("*"):

                    if not archivo.is_file():
                        continue

                    if archivo.name.startswith("~$"):
                        continue

                    if not pertenece_a_carpeta_articulo(archivo, articulo, carpeta_id):
                        continue

                    nombre_normalizado = archivo.stem.strip().upper()

                    if not nombre_normalizado.startswith(prefijo_resumen):
                        continue

                    if archivo.suffix.lower() == ".pdf":
                        pdf_resumen_id.append(str(archivo))

                    elif archivo.suffix.lower() in extensiones_excel:
                        excel_resumen_id.append(str(archivo))

            # ========================================================
            # SI HEMOS ENCONTRADO PDF EN I+D -> TERMINAMOS
            # ========================================================

            if pdf_resumen_id:
                datos_articulo["resumen"] = pdf_resumen_id[0]

                print(f"RESUMEN I+D -> {articulo}: {pdf_resumen_id[0]}")

                continue

            # ========================================================
            # 3. NO EXISTE NINGÚN PDF
            #
            #    Comprobar si al menos existe Excel en alguna
            #    de las dos búsquedas.
            # ========================================================

            excels_encontrados = excel_lotes + excel_resumen_id

            if excels_encontrados:
                datos_articulo["resumen"] = None

                estructura_documental["no_encontrados"].append({
                    "articulo": articulo,
                    "proveedor": "",
                    "tipo": "RESUMEN",
                    "referencia": "LOTES / RESUMEN",
                    "msg": f"Se ha encontrado un Excel de resumen para el artículo {articulo}, pero no existe el PDF correspondiente.",
                    "excel_encontrado": excels_encontrados[0]
                })

                print(f"ERROR RESUMEN -> {articulo}: existe Excel pero falta PDF")

                continue

            # ========================================================
            # 4. NO EXISTE NI PDF NI EXCEL
            # ========================================================

            datos_articulo["resumen"] = None

            estructura_documental["no_encontrados"].append({
                "articulo": articulo,
                "proveedor": "",
                "tipo": "RESUMEN",
                "referencia": "LOTES / RESUMEN",
                "msg": f"No se ha encontrado ningún resumen PDF ni Excel para el artículo {articulo}, ni mediante LOTES en ERP ni mediante {articulo}_RESUMEN en I+D."
            })

            print(f"ERROR RESUMEN -> {articulo}: no existe PDF ni Excel")

        return estructura_documental


    # ========================================================
    # ESTRUCTURA
    # ========================================================

    estructura = {
        "componentes": {},
        "rutas": [],
        "no_encontrados": []
    }

    def obtener_componente(articulo):
        articulo = str(articulo).strip()

        if articulo not in estructura["componentes"]:
            estructura["componentes"][articulo] = {
                "albaran_op": [],
                "resumen": None,
                "pnt": []
            }

        return estructura["componentes"][articulo]

    # ========================================================
    # AÑADIR DOCUMENTO
    # ========================================================

    def añadir_documento(articulo, campo, tipo, referencia, rutas):
        componente = obtener_componente(articulo)
        referencia = normalizar_valor(referencia)

        if referencia is None:
            return

        if tipo in ("ORDEN", "OP"):
            proveedores = ["I+D"]
        else:
            proveedores = []

            for ruta in rutas:
                proveedor = obtener_proveedor_ruta(ruta)

                if proveedor and proveedor not in proveedores:
                    proveedores.append(proveedor)

        for documento in componente[campo]:
            if documento["tipo"] == tipo and documento["referencia"] == referencia:

                for ruta in rutas:
                    if ruta not in documento["rutas"]:
                        documento["rutas"].append(ruta)

                for proveedor in proveedores:
                    if proveedor not in documento["proveedores"]:
                        documento["proveedores"].append(proveedor)

                return

        componente[campo].append({
            "tipo": tipo,
            "referencia": referencia,
            "proveedores": proveedores,
            "rutas": rutas
        })

    # ========================================================
    # NO ENCONTRADOS
    # ========================================================

    def registrar_no_encontrado(articulo, tipo, referencia):
        referencia = normalizar_valor(referencia)

        for existente in estructura["no_encontrados"]:
            if existente["articulo"] == articulo and existente["tipo"] == tipo and existente["referencia"] == referencia:
                return

        proveedor = "I+D" if tipo in ("ORDEN", "OP") else None

        estructura["no_encontrados"].append({
            "articulo": articulo,
            "proveedor": proveedor,
            "tipo": tipo,
            "referencia": referencia,
            "msg": f"No se ha localizado el documento {tipo} '{referencia}' del componente '{articulo}'"
        })

    # ========================================================
    # OP CONOCIDAS
    # ========================================================

    ops_dispositivo = valores_unicos(df_dispositivo, "ORDEN DE PRODUCCION")
    ops_consola = valores_unicos(df_consola, "ORDEN DE PRODUCCION")
    ops_conocidas = set(ops_dispositivo + ops_consola)

    # ========================================================
    # PROCESAR COLUMNAS DE LOS DOS DATAFRAMES
    # ========================================================

    def procesar_dataframe(df):

        columnas_albaran = [columna for columna in df.columns if str(columna).strip().upper().endswith("_ALBARAN")]

        for columna in columnas_albaran:
            articulo = str(columna).strip()[:-len("_ALBARAN")]
            obtener_componente(articulo)

            for referencia in valores_unicos(df, columna):

                if str(referencia).strip().startswith("1//"):
                    tipo = "ORDEN"
                elif referencia in ops_conocidas:
                    tipo = "OP"
                else:
                    tipo = "ALBARAN"

                rutas = buscar_albaran_op(referencia, tipo, articulo)
                añadir_documento(articulo, "albaran_op", tipo, referencia, rutas)

                if not rutas:
                    registrar_no_encontrado(articulo, tipo, referencia)

        columnas_pnt = [columna for columna in df.columns if str(columna).strip().upper().endswith("_PNT")]

        for columna in columnas_pnt:
            articulo = str(columna).strip()[:-len("_PNT")]
            obtener_componente(articulo)

            for referencia in valores_unicos(df, columna):
                rutas = buscar_pnt(referencia)
                añadir_documento(articulo, "pnt", "PNT", referencia, rutas)

                if not rutas:
                    registrar_no_encontrado(articulo, "PNT", referencia)

    procesar_dataframe(df_dispositivo)
    procesar_dataframe(df_consola)

    # ========================================================
    # AÑADIR OP PRINCIPALES
    #
    # Identificamos el artículo buscando qué columna *_ALBARAN
    # contiene esas OP.
    # ========================================================

    def buscar_articulo_de_op(df, op):
        for columna in df.columns:
            if not str(columna).strip().upper().endswith("_ALBARAN"):
                continue

            if op in valores_unicos(df, columna):
                return str(columna).strip()[:-len("_ALBARAN")]

        return None

    for op in ops_dispositivo:
        articulo = buscar_articulo_de_op(df_dispositivo, op)

        if articulo:
            rutas = buscar_albaran_op(op, "OP", articulo)
            añadir_documento(articulo, "albaran_op", "OP", op, rutas)

            if not rutas:
                registrar_no_encontrado(articulo, "OP", op)

    for op in ops_consola:
        articulo = buscar_articulo_de_op(df_consola, op)

        if articulo:
            rutas = buscar_albaran_op(op, "OP", articulo)
            añadir_documento(articulo, "albaran_op", "OP", op, rutas)

            if not rutas:
                registrar_no_encontrado(articulo, "OP", op)

    # ========================================================
    # INFERIR PROVEEDOR DE ALBARANES NO ENCONTRADOS
    #
    # Solo se asigna si el artículo tiene UN ÚNICO proveedor
    # conocido. Nunca se mezclan varios proveedores.
    # ========================================================

    def obtener_proveedor_articulo(articulo):
        componente = estructura["componentes"].get(articulo)

        if componente is None:
            return None

        proveedores = set()

        for documento in componente["albaran_op"]:
            if documento["tipo"] != "ALBARAN":
                continue

            for proveedor in documento["proveedores"]:
                if proveedor:
                    proveedores.add(proveedor)

        if len(proveedores) == 1:
            return next(iter(proveedores))

        return None

    for documento in estructura["no_encontrados"]:
        if documento["tipo"] in ("ORDEN", "OP"):
            documento["proveedor"] = "I+D"
        else:
            documento["proveedor"] = obtener_proveedor_articulo(documento["articulo"]) or "NO IDENTIFICADO"

    # ========================================================
    # LISTA PLANA DE RUTAS
    # ========================================================

    rutas_finales = []
    vistos = set()

    for articulo, componente in estructura["componentes"].items():
        for campo in ("albaran_op", "pnt"):
            for documento in componente[campo]:
                for ruta in documento["rutas"]:

                    clave = (articulo, documento["tipo"], documento["referencia"], ruta)

                    if clave in vistos:
                        continue

                    vistos.add(clave)

                    rutas_finales.append({
                        "articulo": articulo,
                        "tipo": documento["tipo"],
                        "referencia": documento["referencia"],
                        "ruta": ruta
                    })

    estructura["rutas"] = rutas_finales


    estructura = localizar_resumenes_articulos(estructura)

    # ========================================================
    # CONTROL FINAL DE INTEGRIDAD
    #
    # Una ORDEN/OP no puede apuntar a la carpeta I+D de otro
    # artículo.
    # ========================================================

    for articulo, componente in estructura["componentes"].items():
        for documento in componente["albaran_op"]:

            if documento["tipo"] not in ("ORDEN", "OP"):
                continue

            rutas_correctas = []

            for ruta in documento["rutas"]:
                partes = [parte.upper() for parte in Path(ruta).parts]

                if "I+D" not in partes:
                    continue

                indice = partes.index("I+D")

                if indice + 1 < len(partes) and partes[indice + 1] == articulo.upper():
                    rutas_correctas.append(ruta)

            documento["rutas"] = rutas_correctas

    return estructura


def buscar_saltos_numeros_serie(directorio):

    def obtener_ultimos_4(numero_serie):
        if pd.isna(numero_serie):
            return None

        texto = str(numero_serie).strip()

        if texto.endswith(".0"):
            texto = texto[:-2]

        numeros = re.findall(r"\d", texto)

        if len(numeros) < 4:
            return None

        return int("".join(numeros[-4:]))


    def convertir_a_rangos(numeros_faltantes):
        if not numeros_faltantes:
            return []

        rangos = []
        inicio = numeros_faltantes[0]
        anterior = numeros_faltantes[0]

        for numero in numeros_faltantes[1:]:
            if numero == anterior + 1:
                anterior = numero
                continue

            rangos.append((inicio, anterior))
            inicio = numero
            anterior = numero

        rangos.append((inicio, anterior))

        return rangos


    def rangos_a_texto(rangos):
        textos = []

        for inicio, fin in rangos:
            if inicio == fin:
                textos.append(f"{inicio:04d}")
            else:
                textos.append(f"{inicio:04d}-{fin:04d}")

        return ", ".join(textos)


    directorio = Path(directorio)
    resultados = []

    extensiones = {".xlsx", ".xlsm", ".xls"}

    archivos = [archivo for archivo in directorio.rglob("*") if archivo.is_file() and "_RESUMEN" in archivo.stem.upper() and archivo.suffix.lower() in extensiones and not archivo.name.startswith("~$")]

    for archivo in archivos:

        articulo = archivo.parent.name

        try:
            excel = pd.ExcelFile(archivo)
        except Exception as e:
            print(f"No se ha podido abrir {archivo}: {e}")
            continue

        for hoja in excel.sheet_names:

            try:
                df = pd.read_excel(archivo, sheet_name=hoja, dtype=object)
            except Exception as e:
                print(f"No se ha podido leer {archivo} - {hoja}: {e}")
                continue

            columnas = {str(columna).strip().upper(): columna for columna in df.columns}

            if "NUMERO_SERIE" not in columnas:
                continue

            columna_ns = columnas["NUMERO_SERIE"]

            numeros = []

            for numero_serie in df[columna_ns]:
                numero = obtener_ultimos_4(numero_serie)

                if numero is not None:
                    numeros.append(numero)

            numeros = sorted(set(numeros))

            if len(numeros) < 2:
                continue

            numero_inicial = numeros[0]
            numero_final = numeros[-1]

            existentes = set(numeros)
            faltantes = [numero for numero in range(numero_inicial, numero_final + 1) if numero not in existentes]
            rangos = convertir_a_rangos(faltantes)

            if not rangos:
                continue

            resultados.append({
                "articulo": articulo,
                "rango_existente": f"{numero_inicial:04d}-{numero_final:04d}",
                "saltos": rangos_a_texto(rangos),
                "archivo": str(archivo)
            })

    return resultados

DF_PNT = leer_excel_PNTs()
dic_config = leer_configuracion()

if __name__ == "__main__":
        
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ PROCESAR TRAZABILIDAD Y DOCUMENTACION ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━═════════════════════════════════════════════════════════════════════════════┛

    # articulos_IMD = os.listdir(RUTA_IMD)      # Todos los articulos 
    articulos_IMD = ['CABLEPLANO']        # Elegir manualmente los articulos

    articulos_procesados = [] # Lista que contiene los articulos ya procesados
    articulos_pendientes = articulos_IMD.copy()
    print("\n Combirtiendo lotes a PDF")

    consulta = input(f"\nQuieres regenerar los .pdf de los LOTES de materias NO fabricadas Tarda MUCHO? (S/N): ")
    if consulta.capitalize() == "S":
        convertir_excels_lotes_erp_a_pdf() # Recorre todos los excels de ERP menos los de I+D y lso combierte a PDF

    consulta = input(f"\nQuieres procesar las OP?(S/N): ")
    if consulta.capitalize() == "S":

        for i in INDICE_ARTICULOS:

            # print("\nProcesando los articulos de: ",i,'\n')

            # Procesamos los articulos ordenados de menos procesado a mas procesado
            for articulo in dic_config[i]:
                
                # if i  in 'DISPOSITIVOS':
                #     break # No procesamos aun los dispositivos

                # if i in 'MATERIA PRIMA N2':
                #     break # No procesamos aun la materia prima de nivel 2 (Consolas)

                # if i in 'MATERIA PRIMA N1':
                #     break # No procesamos aun la materia prima de nivel 1

                if articulo in articulos_IMD:
                    
                    print("Procesando el articulo: ",articulo)
                    articulos_procesados.append(articulo)
                    articulos_pendientes.remove(articulo)
                    
                    ruta_carpeta_articulo =  Path(RUTA_IMD + articulo +'/')

                    normalizar_nombres_archivos(ruta_carpeta_articulo) # Normalizamos los nombres de los archivos en la carpeta del artículo

                    ordenes_produccion_excel = [
                        archivo
                        for archivo in ruta_carpeta_articulo.iterdir()
                        if archivo.is_file()
                        and archivo.suffix.lower() in (".xlsx", ".xls")
                        and patron_orden_produccion.fullmatch(archivo.stem.strip())
                    ]
                    
                    ordenes_produccion_excel = ordenar_ordenes_produccion(ordenes_produccion_excel)

                    lista_dataframes_ordenes = [] # es una lista que contiene todos los dataframes de consumos de las ordenes que se van a procesar y se usar para unirlos luego al resumen
                    
                    # Leemos el resuemnexistente de ordnes del articulo si existe, sino lo creamos
                    try:
                        df_resumen, ordenes_procesadas, ruta_resumen = leer_excel_resumen(ruta_carpeta_articulo)

                    except:
                        
                        print(f" \nAVISO: No se pudo leer el resumen existente en {ruta_carpeta_articulo} se va a realizar el reseteo del resumen del articulo {articulo}.\n")
                        time.sleep(1)
                        
                        ordenes_procesadas = []
                        df_resumen = pd.DataFrame()
                        ruta_resumen = ruta_carpeta_articulo / f"{articulo}_RESUMEN.xlsx"

                    else:
                        lista_dataframes_ordenes.append(df_resumen)
                    
                    if MODO_RESET: # Fuerza que se procesen todas las carpetas de ordenes desde 0

                        lista_dataframes_ordenes = [] # Limpiamos la lista
                        ordenes_procesadas = []
                        ruta_resumen = ruta_carpeta_articulo / f"{articulo}_RESUMEN.xlsx"
                        df_resumen = pd.DataFrame()
                    
                    ordenes_por_procesar = obtener_ordenes_por_procesar(ordenes_produccion_excel, ordenes_procesadas)
                    # print("\ordenes_por_procesar del articulo {articulo}:\n",ordenes_por_procesar)
            
                    for ruta_orden_excel in ordenes_por_procesar:
                        # print(f"Procesando orden de producción: {ruta_orden_excel.name}...")
                        print("\n")
                        df_consumos_orden = procesar_orden(ruta_orden_excel)

                        if GENERAR_PDF_ORDENES:
                            ruta_pdf = str(Path(ruta_orden_excel).with_suffix(".pdf"))
                            if not Path(ruta_pdf).exists():
                                print(f"Generando PDF de la orden {ruta_orden_excel.name}...")
                                convertir_orden_excel_a_pdf(ruta_orden_excel)
                            else:
                                print(f"El PDF de la orden {ruta_orden_excel.name} ya existe, no se generará de nuevo.")

                        if df_consumos_orden is not None: # Si no es none
                            lista_dataframes_ordenes.append(df_consumos_orden)
                    
                    df_resumen = pd.concat(lista_dataframes_ordenes,ignore_index=True)

                    # Guardamos el dataframe en el excel resumen del artuculo
                    df_resumen.to_excel(ruta_resumen, index=False)

                    ruta_excel_formateado = ruta_resumen.with_name(f"{ruta_resumen.stem}_formateado{ruta_resumen.suffix}")
                    print(f"\nFormateando resumen excel del articulo {articulo} ...")
                    formatear_resumen_excel(ruta_resumen, ruta_salida = ruta_excel_formateado, hoja=None, device = articulo)
                    print(f"\nImprimiendo resumen excel del articulo {articulo} ...")
                    excel_resumen_a_pdf_una_hoja(ruta_excel_formateado)

        # Guardamos el dataframe en el excel resumen del artuculo
        df_errores = pd.DataFrame(set(lista_errores))
        df_errores.to_excel(
            './errores.xlsx',
            index=False
        )
                
        print('\nSe han procesado los articulos:',articulos_procesados)
        print('\nQuedan pendientes de procesar:',articulos_pendientes,'\n')
        time.sleep(2)

        saltos = buscar_saltos_numeros_serie(RUTA_IMD)
        if saltos:
            print("\nSALTOS DE NUMERACIÓN DETECTADOS:\n")

            for resultado in saltos:
                print(f'{resultado["articulo"]}: {resultado["saltos"]}')

            consulta = input(f"\nSe detectaron OP faltantes que causan saltos de numeros de serie que se muestran más arriba , quieres continuar realizando el BachRecord?(S/N): ")
            if consulta.capitalize() != "S":
                raise ValueError(
                    f"Se termino la ejecucion del bachrecord por falta de informacion en en las ordenes de produccion" )
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ PROCESAMOS INFORMACION EXCLUSIVA DEL BACHRECORD ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━═════════════════════════════════════════════════════════════════════════════┛
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    # PRIMER PEDIMOS LOS NUMEROS DE SERIE DEL BACHRECORD QUE SE QUIERE GENERAR
    
    procesar = input("Quiere procesar el bachrecord de un dispositivo? (S/N): ").strip().lower() # TODO DESCOMENTAR AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
    procesar = "s"# TODO BORRAR AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA

    if procesar == "s":

        # dispositivo = input("Introduzca el nombre del dispositivo: ").strip()# TODO DESCOMENTAR AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
        dispositivo = "EPTEV02DEV01"# TODO BORRAR  AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA

        while dispositivo not in DISPOSITIVOS:
            print(f"El dispositivo '{dispositivo}' no es válido. Los dispositivos : {', '.join(DISPOSITIVOS)}")
            dispositivo = input("Introduzca el nombre del dispositivo: ").strip()

        # Obtenemos el DataFrame del resumen de la orden de producción del dispositivo
        ruta_carpeta_articulo =  Path(RUTA_IMD + dispositivo +'/')
        df_resumen, _, _ = leer_excel_resumen(ruta_carpeta_articulo)
        ns_producidos = (df_resumen["NUMERO_SERIE"].dropna().astype(str).unique().tolist())
        
        ns_inicio_br = "EPB1250384" # TODO BORRAR  AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
        ns_final_br = "EPB1260953" # TODO BORRAR  AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA

        # ns_inicio_br = input("Introduzca el número de serie inicial: ").strip()
        # while ns_inicio_br not in ns_producidos:
        #     print(f"El número de serie '{ns_inicio_br}' no es válido. Los números de serie estan comprendidos entre: {ns_producidos[0]} y {ns_producidos[-1]}")
        #     ns_inicio_br = input("Introduzca el número de serie inicial: ").strip()

        # ns_final_br = input("Introduzca el número de serie final: ").strip()
        # while ns_final_br not in ns_producidos or ns_final_br < ns_inicio_br:
        #     if ns_final_br < ns_inicio_br:
        #         print(f"El número de serie final '{ns_final_br}' no puede ser menor que el número de serie inicial '{ns_inicio_br}'.")
        #     elif ns_final_br not in ns_producidos:
        #         print(f"El número de serie '{ns_final_br}' no es válido. Los números de serie estan comprendidos entre: {ns_producidos[0]} y {ns_producidos[-1]}")
        #     ns_final_br = input("Introduzca el número de serie final: ").strip()

        # A PARTIR DE ESTE PUNTO YA TENEMOS EL DISPOSITIVO Y LOS NUMEROS DE SERIE QUE SE QUIEREN PROCESAR PARA EL BACHRECORD
        df_dispoitivoBR = filtrar_filas_entre_valores(df_resumen, "NUMERO_SERIE", ns_inicio_br, ns_final_br) # Nos quedamos unicamente con los ns que queremos para el bachredord
        df_dispoitivoBR = eliminar_articulos_vacios_df(df_dispoitivoBR) # Eliminamos los articulos vacios que se generan por culpa de haber cambiado las estructuras de las producciones
        df_dispoitivoBR.to_excel('./temp/DispositivoBR_temp.xlsx', index=False)# guardamos en excel si queremos consultar algo

        info_faltante = revisa_informacion_bachrecord(df_dispoitivoBR, dispositivo)

        if info_faltante:
            for error in info_faltante:
                print(error["msg"])

            consulta = input(f"\nSe detecto informacion faltante en el resumen principal puedes revisar la informacion en './temp/DispositivoBR_temp.xlsx', quieres continuar realizando el BachRecord? (S/N): ")
            if consulta.capitalize() != "S":
                raise ValueError(
                    f"Se termino la ejecucion del bachrecord por falta de informacion en './temp/DispositivoBR_temp.xlsx'" )

        # Antes de formatear el excel nos hemos aseguraamos de que todo el contenido que necesitmaos esta
        formatear_resumen_excel('./temp/DispositivoBR_temp.xlsx', ruta_salida='./temp/DispTempFiltrado.xlsx', hoja=None, device = dispositivo)
        print("Archivo temporal filtrado y formateado guardado en './temp/DispTempFiltrado.xlsx'.")

        # REPRETIMOS EL PROCESO ANTERIOR CON LA CONSOLA
        consola = DISPOSITIVO_CONSOLA[dispositivo]

        ruta_excel_consola =  RUTA_IMD + consola +"/"+ f"{consola}_RESUMEN.xlsx" 
        df_consola = pd.read_excel(ruta_excel_consola)

        df_consolaBR = filtrar_filas_entre_valores(df_consola, "NUMERO_SERIE", ns_inicio_br, ns_final_br) # Nos quedamos unicamente con los ns que queremos para el bachredord
        df_consolaBR = eliminar_articulos_vacios_df(df_consolaBR) # Eliminamos los articulos vacios que se generan por culpa de haber cambiado las estructuras de las producciones
        df_consolaBR.to_excel('./temp/ConsolaBR_temp.xlsx', index=False) # Guardamos en excel si queremos consultar algo

        info_faltante = revisa_informacion_bachrecord(df_consolaBR, consola)

        if info_faltante:
            for error in info_faltante:
                print(error["msg"])

            consulta = input(f"\nSe detecto informacion faltante en el resumen principal de la consola puedes revisar la informacion en './temp/ConsolaBR_temp.xlsx', quieres continuar realizando el BachRecord? (S/N): ")
            if consulta.capitalize() != "S":
                raise ValueError(
                    f"Se termino la ejecucion del bachrecord por falta de informacion en el resumen principal de la consola './temp/ConsolaBR_temp.xlsx' " )


        # Antes de formatear el excel nos aseguraamos de que todo el contenido que necesitmaos esta
        formatear_resumen_excel('./temp/ConsolaBR_temp.xlsx', ruta_salida='./temp/ConsolaTempFiltrado.xlsx', hoja=None, device = consola)
        print("Archivo temporal filtrado y formateado guardado en './temp/ConsolaTempFiltrado.xlsx'.")

        estructura_documental = localizar_documentacion_batchrecord(dispositivo_br=df_dispoitivoBR,consola_br=df_consolaBR,)

            
        if estructura_documental["no_encontrados"]:

            print("\nHAY DOCUMENTOS NO ENCONTRADOS:")

            for documento in estructura_documental["no_encontrados"]:
                print(f'{documento["articulo"]} - {documento["proveedor"]} - {documento["tipo"]} - {documento["referencia"]}')

        else:

            print("Se ha localizado toda la documentación.")

        guardar_diccionario(estructura_documental)
        print("Estructura documental guardada correctamente")

    estructura_documental = cargar_diccionario()

    for a in estructura_documental['componentes']:
        b = estructura_documental['componentes'][a]["resumen"]
        print(b)
        # for a in estructura_documental['componentes']


    # for valor in estructura_documental["componentes"]["PHB1V01"]['albaran_op']:

    #     # pprint(clave)
    #     pprint(valor['rutas'])
    #     time.sleep(1)

    # print("\n" + "=" * 100)
    # print("ERRORES DE ESTRUCTURA DOCUMENTAL")
    # print("=" * 100)

    # errores = estructura_documental.get("no_encontrados", [])

    # if not errores:
    #     print("No se han encontrado errores.")
    # else:
    #     for i, error in enumerate(errores, start=1):
    #         print(f"\nERROR {i}")
    #         print("-" * 100)

    #         for clave, valor in error.items():
    #             print(f"{clave.upper()}: {valor}")

    # print("\n" + "=" * 100)
    # print(f"TOTAL ERRORES: {len(errores)}")
    # print("=" * 100)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ CREAR PDF ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━═════════════════════════════════════════════════════════════════════════════┛
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    # rec = obtener_requisitos_especiales("EPTEV02DEV02","0000","0575")
    # bachredord = GeneradorBatchRecord("BATCH_RECORD_EPTE.pdf")
    # bachredord.page_crear_portada(

    #         lote="XXXX",

    #         dispositivo="EPTEV02DEV02",

    #         ns_inicio="EPB1230001",
    #         ns_final="EPB1230090",

    #         software_version="4643_3385",

    #         requisitos_especiales=rec,

    #         preparado_por="Victoria E. González Gutiérrez",
    #         cargo_preparado="Regulatory Responsible",

    #         revisado_por="Victoria E. González Gutiérrez",
    #         cargo_revisado="Quality Manager",

    #         aprobado_por="Josep Oliver Garcia",
    #         cargo_aprobado="Manager",

    #         fecha_preparado="2024/02/15",
    #         fecha_revisado="2024/02/15",
    #         fecha_aprobado="2024/02/15",
    #     )
    # bachredord.page_crear_indice()
    # bachredord.guardar()